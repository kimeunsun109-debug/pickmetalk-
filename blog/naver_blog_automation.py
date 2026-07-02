#!/usr/bin/env python3
"""
네이버 블로그 이웃글 순회 자동화
- Chrome CDP(9222) 연결 — 저장된 아이디/비밀번호(자동완성) 로그인
- 방문 기록 JSONL 저장 (같은 블로거 당일 재방문 금지)
- 이웃댓글_상황별문장.xlsx 에서 상황 맞는 댓글 선택

사용법:
  1) start_chrome_for_blog.bat 실행 후 네이버 로그인(최초 1회)
  2) python naver_blog_automation.py --mode probe   # UI 탐색
  3) python naver_blog_automation.py --mode visit --limit 3
"""

from __future__ import annotations

import argparse
import json
import random
import re
import sys
import time
import urllib.parse
from datetime import date, datetime
from pathlib import Path

if hasattr(sys.stdout, "reconfigure"):
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

import openpyxl
from playwright.sync_api import Page, TimeoutError as PWTimeout, sync_playwright

BASE = Path(__file__).resolve().parent
CDP = "http://127.0.0.1:9222"
MY_BLOG = "https://blog.naver.com/hsrw1011"
NEIGHBOR_HOME = (
    "https://section.blog.naver.com/BlogHome.naver"
    "?directoryNo=0&currentPage=1&groupId=0"
)
NEIGHBOR_POSTS = (
    "https://section.blog.naver.com/connect/ViewMoreBuddyPosts.naver"
    "?blogId=hsrw1011&currentPage=1"
)
SKIP_BLOG_IDS = {
    "BlogHome.naver", "MyBlog.naver", "market", "blogpeople",
    "section", "hsrw1011", "PostList.naver", "BlogHome",
}
LOGIN_URL = "https://nid.naver.com/nidlogin.login"
VISIT_LOG = BASE / "naver_visit_log.jsonl"
COMMENTS_XLSX = BASE / "이웃댓글_상황별문장.xlsx"
STAY_SEC = 30


def log_visit(entry: dict) -> None:
    entry.setdefault("ts", datetime.now().isoformat(timespec="seconds"))
    with VISIT_LOG.open("a", encoding="utf-8") as f:
        f.write(json.dumps(entry, ensure_ascii=False) + "\n")


def load_visited_today() -> set[str]:
    today = date.today().isoformat()
    seen: set[str] = set()
    if not VISIT_LOG.exists():
        return seen
    for line in VISIT_LOG.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        try:
            e = json.loads(line)
        except json.JSONDecodeError:
            continue
        if e.get("date") == today and e.get("blog_id"):
            seen.add(e["blog_id"])
    return seen


def load_comments() -> list[dict]:
    if not COMMENTS_XLSX.exists():
        raise FileNotFoundError(f"댓글 엑셀 없음: {COMMENTS_XLSX}")
    wb = openpyxl.load_workbook(COMMENTS_XLSX, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for r in range(2, ws.max_row + 1):
        situation = str(ws.cell(r, 1).value or "").strip()
        text = str(ws.cell(r, 2).value or "").strip()
        hints = str(ws.cell(r, 3).value or "").strip()
        if text:
            rows.append({"situation": situation, "text": text, "hints": hints})
    return rows


def pick_comment(title: str, body_snippet: str, pool: list[dict]) -> str:
    blob = f"{title} {body_snippet}".lower()
    scored: list[tuple[int, dict]] = []
    for item in pool:
        score = 0
        for hint in re.split(r"[,，]", item.get("hints", "")):
            hint = hint.strip().lower()
            if hint and hint in blob:
                score += 2
        if item["situation"].startswith("마무리") and score == 0:
            score = 1
        scored.append((score, item))
    scored.sort(key=lambda x: x[0], reverse=True)
    top_score = scored[0][0] if scored else 0
    candidates = [it for sc, it in scored if sc == top_score] or pool
    # 질문형은 20% 확률로만
    non_question = [c for c in candidates if c["situation"] != "질문형"]
    if non_question and random.random() > 0.2:
        candidates = non_question
    return random.choice(candidates)["text"]


def get_cdp_page(browser, prefer: str | None = None) -> Page:
    for ctx in browser.contexts:
        for pg in ctx.pages:
            if prefer and prefer in pg.url:
                return pg
    for ctx in browser.contexts:
        if ctx.pages:
            return ctx.pages[0]
    ctx = browser.contexts[0] if browser.contexts else browser.new_context()
    return ctx.new_page()


def is_logged_in(page: Page) -> bool:
    try:
        page.goto(MY_BLOG, wait_until="domcontentloaded", timeout=20000)
        page.wait_for_timeout(2000)
        url = page.url
        if "nidlogin" in url or "nid.naver.com" in url:
            return False
        # 내 블로그 접근 + 로그인 UI 없으면 성공으로 간주
        if "hsrw1011" in url or "blog.naver.com" in url:
            if page.locator("text=로그인").count() == 0:
                return True
            # 글쓰기 버튼 등
            if page.locator("a:has-text('글쓰기'), button:has-text('글쓰기')").count():
                return True
        return "nidlogin" not in page.url
    except PWTimeout:
        return False


def login_with_saved_credentials(page: Page) -> bool:
    """저장된 아이디/비밀번호 — 네이버 최근계정 클릭 또는 브라우저 자동완성."""
    page.goto(LOGIN_URL, wait_until="domcontentloaded", timeout=20000)
    page.wait_for_timeout(1500)

    # 1) 네이버 최근 로그인 계정(원클릭) — UI 변경 대비 여러 셀렉터
    quick_selectors = [
        "#rcapt li a",
        ".recommend_account a",
        "a[href*='nidlogin.login'][class*='account']",
        ".login_box .profile",
        "button:has-text('로그인')",
        ".btn_login_wrap a",
    ]
    for sel in quick_selectors:
        loc = page.locator(sel)
        if loc.count() > 0:
            try:
                loc.first.click(timeout=3000)
                page.wait_for_timeout(2500)
                if "nidlogin" not in page.url:
                    return True
            except Exception:
                pass

    # 2) 아이디 칸 클릭 → 브라우저 저장 비밀번호 자동완성 유도
    id_box = page.locator("#id, input[name='id']").first
    pw_box = page.locator("#pw, input[name='pw']").first
    if id_box.count():
        id_box.click()
        page.wait_for_timeout(800)
        # 키보드로 자동완성 선택 시도 (Chrome: 아래 화살표 + Enter)
        page.keyboard.press("ArrowDown")
        page.wait_for_timeout(300)
        page.keyboard.press("Enter")
        page.wait_for_timeout(500)
        if pw_box.count():
            pw_box.click()
            page.wait_for_timeout(500)
            page.keyboard.press("ArrowDown")
            page.wait_for_timeout(300)
            page.keyboard.press("Enter")
            page.wait_for_timeout(500)

        login_btn = page.locator("#log\\.login, .btn_login, button:has-text('로그인')").first
        if login_btn.count():
            login_btn.click()
            page.wait_for_timeout(3000)

    if "nidlogin" not in page.url:
        return True

    # 3) 이미 로그인된 다른 탭에서 세션 공유된 경우
    return is_logged_in(page)


def ensure_login(page: Page) -> None:
    if is_logged_in(page):
        print("[로그인] 세션 유지 중")
        return
    print("[로그인] 저장된 계정/자동완성으로 시도…")
    if not login_with_saved_credentials(page):
        raise RuntimeError(
            "로그인 실패. start_chrome_for_blog.bat 으로 Chrome 연 뒤 "
            "네이버에 수동 로그인(아이디 저장 체크) 후 다시 실행하세요."
        )
    print("[로그인] 성공")


def open_neighbor_it_feed(page: Page) -> None:
    page.goto(NEIGHBOR_HOME, wait_until="domcontentloaded", timeout=25000)
    page.wait_for_timeout(2500)

    for label in ("이웃새글", "이웃 새글"):
        tab = page.locator(f"a:has-text('{label}'), button:has-text('{label}'), span:has-text('{label}')")
        if tab.count():
            tab.first.click()
            page.wait_for_timeout(2000)
            break

    for label in ("전체이웃", "전체 이웃"):
        btn = page.locator(f"a:has-text('{label}'), button:has-text('{label}'), span:has-text('{label}')")
        if btn.count():
            btn.first.click()
            page.wait_for_timeout(1500)
            break

    for label in ("IT", "IT·컴퓨터", "IT관련", "IT/컴퓨터", "IT・컴퓨터"):
        it = page.locator(f"a:has-text('{label}'), button:has-text('{label}'), span:has-text('{label}')")
        if it.count():
            it.first.click()
            page.wait_for_timeout(2000)
            break

    # 이웃새글 목록 직접 URL (탭 클릭 실패 시 백업)
    if page.locator("a[href*='logNo='], a[href*='PostView']").count() < 3:
        page.goto(NEIGHBOR_POSTS, wait_until="domcontentloaded", timeout=25000)
        page.wait_for_timeout(2000)


def extract_blog_id(url: str) -> str:
    m = re.search(r"blog\.naver\.com/([^/?#]+)/\d+", url)
    if m:
        return m.group(1)
    m = re.search(r"blog\.naver\.com/([^/?#]+)", url)
    if m and m.group(1) not in ("GoRepresentBlog.naver", "PostView.naver"):
        return m.group(1)
    q = urllib.parse.parse_qs(urllib.parse.urlparse(url).query)
    if "blogId" in q:
        return q["blogId"][0]
    return ""


def collect_post_links(page: Page, limit: int = 10, it_only: bool = True) -> list[dict]:
    """이웃커넥트 목록에서 글 링크 수집 (blog.naver.com/아이디/글번호)."""
    raw: list[dict] = page.evaluate(
        """() => {
        const byUrl = new Map();
        document.querySelectorAll('a[href*="blog.naver.com/"]').forEach(a => {
          const h = a.href.split('#')[0];
          if (!/blog\\.naver\\.com\\/[^/]+\\/\\d+/.test(h)) return;
          const t = (a.innerText || '').trim().replace(/\\s+/g, ' ');
          if (!t || t.length < 4) return;
          if (t.includes('쿠팡 파트너스') || t.startsWith('안녕하')) return;
          const prev = byUrl.get(h);
          if (!prev || (t.length < 80 && t.length > (prev.title || '').length)) {
            byUrl.set(h, { url: h, title: t.slice(0, 120) });
          }
        });
        return Array.from(byUrl.values());
    }"""
    )

    it_kw = re.compile(
        r"IT|아이폰|갤럭시|스마트|가전|노트북|PC|앱|AI|IoT|리뷰|테크|컴퓨터|"
        r"모니터|청소기|에어컨|로봇|허브|개발|코딩|윈도우|맥북|삼성|LG|네이버|"
        r"RGB|LED|백라이트|디지털|전자|기기|서비스|개편",
        re.I,
    )

    links: list[dict] = []
    seen_ids: set[str] = set()
    for item in raw:
        url = item["url"]
        blog_id = extract_blog_id(url)
        if not blog_id or blog_id in SKIP_BLOG_IDS:
            continue
        if blog_id in seen_ids:
            continue
        title = item.get("title") or ""
        if it_only and title and len(title) > 8 and not it_kw.search(title):
            continue
        seen_ids.add(blog_id)
        if not title or len(title) < 4 or title.startswith("안녕"):
            title = f"{blog_id} 글"
        links.append({"url": url, "title": title[:120], "blog_id": blog_id})
        if len(links) >= limit:
            break
    return links


def _blog_frame(page: Page):
    """네이버 블로그 본문은 mainFrame iframe 안에 있음."""
    page.wait_for_timeout(1500)
    fr = page.frame_locator("iframe#mainFrame")
    if fr.locator("body").count():
        return fr
    return page


def like_post(page: Page) -> bool:
    for root in (_blog_frame(page), page):
        for sel in (
            "a.u_likeit_list_btn._button",
            "span.u_likeit_list_btn._button",
            ".u_likeit_list_btn",
            "button:has-text('공감')",
        ):
            loc = root.locator(sel)
            if loc.count():
                try:
                    loc.first.click(timeout=3000)
                    page.wait_for_timeout(800)
                    return True
                except Exception:
                    pass
    return False


def write_comment(page: Page, text: str) -> bool:
    root = _blog_frame(page)
    for sel in (
        "textarea.textarea_box",
        "textarea#naverComment__write_textarea",
        "div.u_cbox_write_area textarea",
        "textarea[placeholder*='댓글']",
    ):
        loc = root.locator(sel)
        if loc.count():
            try:
                loc.first.click(timeout=3000)
                loc.first.fill(text)
                page.wait_for_timeout(500)
                for btn_sel in (
                    "button:has-text('등록')",
                    "a:has-text('등록')",
                    ".u_cbox_btn_upload",
                ):
                    btn = root.locator(btn_sel)
                    if btn.count():
                        btn.first.click(timeout=3000)
                        page.wait_for_timeout(1500)
                        return True
            except Exception:
                pass
    return False


def read_post_snippet(page: Page) -> str:
    root = _blog_frame(page)
    for sel in ("div.se-main-container", "#postViewArea", ".se-component-content"):
        loc = root.locator(sel)
        if loc.count():
            try:
                return loc.first.inner_text(timeout=5000)[:400]
            except Exception:
                pass
    return ""


def visit_one_post(page: Page, post: dict, comment_pool: list[dict], visited: set[str]) -> str:
    blog_id = post["blog_id"]
    if blog_id in visited:
        return "skip_duplicate_today"

    page.goto(post["url"], wait_until="domcontentloaded", timeout=25000)
    page.wait_for_timeout(2000)

    title = post.get("title", "")
    body_snip = read_post_snippet(page)

    print(f"  체류 {STAY_SEC}초…")
    page.wait_for_timeout(STAY_SEC * 1000)

    comment = pick_comment(title, body_snip, comment_pool)
    liked = like_post(page)
    commented = write_comment(page, comment)

    log_visit({
        "date": date.today().isoformat(),
        "blog_id": blog_id,
        "post_url": post["url"],
        "post_title": title,
        "comment": comment if commented else "",
        "liked": liked,
        "commented": commented,
        "action": "visit",
    })
    visited.add(blog_id)
    return "ok" if commented or liked else "partial"


def run_probe(page: Page) -> None:
    ensure_login(page)
    open_neighbor_it_feed(page)
    print(f"[probe] URL: {page.url}")
    posts = collect_post_links(page, limit=5)
    print(f"[probe] 포스트 링크 {len(posts)}개")
    for p in posts:
        print(f"  - {p['blog_id']}: {p['title'][:50]}")
    shot = BASE / "probe_neighbor_feed.png"
    page.screenshot(path=str(shot), full_page=True)
    print(f"[probe] 스크린샷: {shot}")


def run_visit(page: Page, limit: int) -> None:
    ensure_login(page)
    comments = load_comments()
    visited = load_visited_today()
    print(f"[방문] 오늘 이미 방문한 블로거: {len(visited)}명")

    open_neighbor_it_feed(page)
    posts = collect_post_links(page, limit=limit * 3)
    done = 0
    for post in posts:
        if done >= limit:
            break
        if post["blog_id"] in visited:
            print(f"[skip] 오늘 이미 방문: {post['blog_id']}")
            continue
        print(f"[방문] {post['blog_id']} — {post['title'][:40]}")
        result = visit_one_post(page, post, comments, visited)
        print(f"  → {result}")
        if result.startswith("ok") or result == "partial":
            done += 1
        time.sleep(random.uniform(2, 5))
    print(f"[완료] 오늘 처리: {done}건 | 로그: {VISIT_LOG}")


def main():
    ap = argparse.ArgumentParser(description="네이버 블로그 이웃글 순회")
    ap.add_argument("--mode", choices=["probe", "visit"], default="probe")
    ap.add_argument("--limit", type=int, default=3, help="visit 모드에서 처리할 글 수")
    args = ap.parse_args()

    with sync_playwright() as p:
        browser = p.chromium.connect_over_cdp(CDP)
        page = get_cdp_page(browser)
        if args.mode == "probe":
            run_probe(page)
        else:
            run_visit(page, args.limit)


if __name__ == "__main__":
    main()
