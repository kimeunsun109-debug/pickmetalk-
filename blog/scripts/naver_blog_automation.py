#!/usr/bin/env python3
"""네이버 블로그 자동화 — 이웃 교류 / 이웃신청 / 글 임시저장"""

from __future__ import annotations

import argparse
import json
import random
import re
import sys
import time
import urllib.parse
from datetime import date, datetime
from pathlib import Path

if hasattr(sys.stdout, "reconfigure"):
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

import openpyxl
from playwright.sync_api import Page, TimeoutError as PWTimeout, sync_playwright

from naver_comment import CommentWriter, write_neighbor_comment

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"
LOGS = ROOT / "logs"
POSTS = ROOT / "posts"
IMAGES = ROOT / "images"

CDP = "http://127.0.0.1:9222"
MY_BLOG_ID = "hsrw1011"
MY_BLOG = f"https://blog.naver.com/{MY_BLOG_ID}"
NEIGHBOR_POSTS = (
    f"https://section.blog.naver.com/connect/ViewMoreBuddyPosts.naver"
    f"?blogId={MY_BLOG_ID}&currentPage=1"
)
LOGIN_URL = "https://nid.naver.com/nidlogin.login"
WRITE_URL = f"https://blog.naver.com/PostWriteForm.naver?blogId={MY_BLOG_ID}&Redirect=Write&"
VISIT_LOG = LOGS / "naver_visit_log.jsonl"
COMMENT_LOG = LOGS / "naver_comment_log.jsonl"
APPLY_LOG = LOGS / "naver_neighbor_apply_log.jsonl"
COMMENTS_XLSX = DATA / "이웃댓글_상황별문장.xlsx"
BUDDY_XLSX = DATA / "이웃분류_교류관리.xlsx"
STAY_SEC = 30
MAX_NEIGHBORS = 5000

SKIP_BLOG_IDS = {
    "BlogHome.naver", "MyBlog.naver", "market", "blogpeople",
    "section", MY_BLOG_ID, "PostList.naver", "BlogHome",
}

THEME_DIRS = {
    "IT": "https://section.blog.naver.com/ThemePost.naver?directoryNo=19&activeDirectorySeq=0&currentPage=1",
    "AI": "https://section.blog.naver.com/ThemePost.naver?directoryNo=19&activeDirectorySeq=0&currentPage=1",
    "경제": "https://section.blog.naver.com/ThemePost.naver?directoryNo=33&activeDirectorySeq=0&currentPage=1",
    "생활": "https://section.blog.naver.com/ThemePost.naver?directoryNo=15&activeDirectorySeq=0&currentPage=1",
    "인테리어": "https://section.blog.naver.com/ThemePost.naver?directoryNo=14&activeDirectorySeq=0&currentPage=1",
}

REPLY_LOG = LOGS / "naver_reply_log.jsonl"

REPLY_MESSAGES = [
    "와 댓글 감사해요~ SuN도 잘 보고 갈게요 ㅎㅎ 좋은 하루 보내세요!",
    "소중한 댓글 고마워요 😊 다음에도 편하게 들러주세요~",
    "댓글 남겨주셔서 감사해요! SuN 블로그 또 놀러오세요 ㅎㅎ",
    "와 공감해주셔서 힘이 나요~ 좋은 일만 가득하세요!",
    "댓글 감사합니다! 저도 이웃님 글 잘 보고 있어요 😊",
]

GROUP_LABELS = {
    "it": "IT 관련",
    "plain": "그냥이웃",
    "retech": "재테크",
    "daily": "일상글",
}


def _append_jsonl(path: Path, entry: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    entry.setdefault("ts", datetime.now().isoformat(timespec="seconds"))
    with path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(entry, ensure_ascii=False) + "\n")


def load_visited_today() -> set[str]:
    today = date.today().isoformat()
    seen: set[str] = set()
    if not VISIT_LOG.exists():
        return seen
    for line in VISIT_LOG.read_text(encoding="utf-8").splitlines():
        try:
            e = json.loads(line)
        except json.JSONDecodeError:
            continue
        if e.get("date") == today and e.get("blog_id"):
            seen.add(e["blog_id"])
    return seen


def load_visited_all_time() -> set[str]:
    """당일뿐 아니라 오늘 세션에서 중복 방지용 — 당일만 적용."""
    return load_visited_today()


def load_applied_ids() -> set[str]:
    seen: set[str] = set()
    if not APPLY_LOG.exists():
        return seen
    for line in APPLY_LOG.read_text(encoding="utf-8").splitlines():
        try:
            e = json.loads(line)
        except json.JSONDecodeError:
            continue
        if e.get("blog_id"):
            seen.add(e["blog_id"])
    return seen


def load_comments() -> list[dict]:
    wb = openpyxl.load_workbook(COMMENTS_XLSX, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for r in range(2, ws.max_row + 1):
        situation = str(ws.cell(r, 1).value or "").strip()
        text = str(ws.cell(r, 2).value or "").strip()
        hints = str(ws.cell(r, 3).value or "").strip() if ws.cell(r, 3).value else ""
        if not text:
            continue
        if situation.startswith("랜덤"):
            parts = re.split(r"\s*/\s*", text)
            for part in parts:
                part = part.strip()
                if part:
                    rows.append({"situation": "랜덤", "text": part, "hints": ""})
        else:
            rows.append({"situation": situation, "text": text, "hints": hints})
    return rows


def pick_comment(title: str, body_snippet: str, pool: list[dict]) -> str:
    blob = f"{title} {body_snippet}"
    scored: list[tuple[int, dict]] = []
    for item in pool:
        if item["situation"] == "랜덤":
            continue
        score = 0
        for hint in re.split(r"[,，]", item.get("hints", "")):
            hint = hint.strip()
            if hint and hint in blob:
                score += 2
        if item["situation"].startswith("마무리") and score == 0:
            score = 1
        scored.append((score, item))
    scored.sort(key=lambda x: x[0], reverse=True)
    top = scored[0][0] if scored else 0
    candidates = [it for sc, it in scored if sc == top] or pool
    non_q = [c for c in candidates if c["situation"] != "질문형"]
    if non_q and random.random() > 0.2:
        candidates = non_q
    if top == 0 and random.random() < 0.15:
        randoms = [c for c in pool if c["situation"] == "랜덤"]
        if randoms:
            return random.choice(randoms)["text"]
    return random.choice(candidates)["text"]


def get_cdp_page(browser) -> Page:
    for ctx in browser.contexts:
        if ctx.pages:
            page = ctx.pages[0]
            break
    else:
        page = browser.contexts[0].new_page()

    def _dismiss_dialog(dialog):
        try:
            dialog.dismiss()
        except Exception:
            pass

    try:
        page.on("dialog", _dismiss_dialog)
    except Exception:
        pass
    return page


def is_logged_in(page: Page) -> bool:
    try:
        page.goto(MY_BLOG, wait_until="domcontentloaded", timeout=20000)
        page.wait_for_timeout(2000)
        return "nidlogin" not in page.url
    except PWTimeout:
        return False


def login_with_saved_credentials(page: Page) -> bool:
    page.goto(LOGIN_URL, wait_until="domcontentloaded", timeout=20000)
    page.wait_for_timeout(1500)
    for sel in ("#rcapt li a", ".recommend_account a", ".login_box .profile"):
        loc = page.locator(sel)
        if loc.count():
            try:
                loc.first.click(timeout=3000)
                page.wait_for_timeout(2500)
                if "nidlogin" not in page.url:
                    return True
            except Exception:
                pass
    id_box = page.locator("#id, input[name='id']").first
    pw_box = page.locator("#pw, input[name='pw']").first
    if id_box.count():
        id_box.click()
        page.wait_for_timeout(500)
        page.keyboard.press("ArrowDown")
        page.keyboard.press("Enter")
        if pw_box.count():
            pw_box.click()
            page.wait_for_timeout(400)
            page.keyboard.press("ArrowDown")
            page.keyboard.press("Enter")
        btn = page.locator("#log\\.login, .btn_login").first
        if btn.count():
            btn.click()
            page.wait_for_timeout(3000)
    return "nidlogin" not in page.url or is_logged_in(page)


def ensure_login(page: Page) -> None:
    if is_logged_in(page):
        print("[로그인] 세션 유지")
        return
    if not login_with_saved_credentials(page):
        raise RuntimeError("로그인 실패 — Chrome(9222)에서 네이버 로그인 후 재시도")
    print("[로그인] 성공")


def select_neighbor_group(page: Page, group_key: str = "it") -> None:
    page.goto(NEIGHBOR_POSTS, wait_until="domcontentloaded", timeout=25000)
    page.wait_for_timeout(2000)
    label = GROUP_LABELS.get(group_key, "IT 관련")
    sel = page.locator("select").first
    if sel.count():
        try:
            sel.select_option(label=label)
            page.wait_for_timeout(2000)
            print(f"[그룹] '{label}' 선택")
            return
        except Exception:
            pass
    for opt in page.locator("select option").all():
        t = (opt.inner_text() or "").strip()
        if label in t or t in label:
            val = opt.get_attribute("value")
            if val:
                sel.select_option(value=val)
                page.wait_for_timeout(2000)
                print(f"[그룹] '{t}' 선택")
                return
    print(f"[그룹] '{label}' 선택 실패 — 전체 목록 사용")


def extract_blog_id(url: str) -> str:
    m = re.search(r"blog\.naver\.com/([^/?#]+)/\d+", url)
    if m:
        return m.group(1)
    m = re.search(r"blog\.naver\.com/([^/?#]+)", url)
    if m and m.group(1) not in ("GoRepresentBlog.naver", "PostView.naver"):
        return m.group(1)
    return ""


def collect_post_links(page: Page, limit: int = 20, it_only: bool = False) -> list[dict]:
    raw: list[dict] = page.evaluate(
        """() => {
        const byUrl = new Map();
        document.querySelectorAll('a[href*="blog.naver.com/"]').forEach(a => {
          const h = a.href.split('#')[0];
          if (!/blog\\.naver\\.com\\/[^/]+\\/\\d+/.test(h)) return;
          const t = (a.innerText || '').trim().replace(/\\s+/g, ' ');
          if (!t || t.length < 4) return;
          if (t.includes('쿠팡 파트너스')) return;
          const prev = byUrl.get(h);
          if (!prev || (t.length < 80 && t.length > (prev.title || '').length)) {
            byUrl.set(h, { url: h, title: t.slice(0, 120) });
          }
        });
        return Array.from(byUrl.values());
    }"""
    )
    links: list[dict] = []
    seen_ids: set[str] = set()
    for item in raw:
        url = item["url"]
        blog_id = extract_blog_id(url)
        if not blog_id or blog_id in SKIP_BLOG_IDS or blog_id in seen_ids:
            continue
        title = item.get("title") or f"{blog_id} 글"
        seen_ids.add(blog_id)
        links.append({"url": url, "title": title[:120], "blog_id": blog_id})
        if len(links) >= limit:
            break
    return links


def _main_frame(page: Page):
    page.wait_for_selector("iframe#mainFrame", timeout=15000)
    return page.frame_locator("iframe#mainFrame")


def like_post(page: Page) -> bool:
    try:
        fl = _main_frame(page)
        result = fl.locator("body").evaluate("""() => {
          const face = document.querySelector('a.u_likeit_button._face.off, a.u_likeit_button._face');
          if (face) { face.scrollIntoView({block:'center'}); face.click(); }
          const like = document.querySelector('a.u_likeit_list_button[data-type="like"]');
          if (like) { like.click(); return true; }
          return !!face;
        }""")
        page.wait_for_timeout(800)
        return bool(result)
    except Exception:
        return False


def write_comment(page: Page, text: str, *, post_url: str = "", title: str = "") -> bool:
    """이웃글 댓글 작성 — naver_comment 모듈 위임."""
    url = post_url or page.url
    writer = CommentWriter(page, log_path=COMMENT_LOG)
    result = writer.write(url, text, title=title)
    if result.ok:
        print(f"  [댓글] 성공 ({result.strategy})")
    else:
        print(f"  [댓글] 실패 — {result.error or 'unknown'}")
    return result.ok


def read_post_snippet(page: Page) -> str:
    try:
        fl = _main_frame(page)
        for sel in ("div.se-main-container", "#postViewArea"):
            loc = fl.locator(sel)
            if loc.count():
                return loc.first.inner_text(timeout=5000)[:400]
    except Exception:
        pass
    return ""


def visit_one_post(page: Page, post: dict, comment_pool: list[dict], visited: set[str], group: str) -> str:
    blog_id = post["blog_id"]
    if blog_id in visited:
        return "skip_duplicate_today"
    try:
        page.goto(post["url"], wait_until="domcontentloaded", timeout=25000)
    except Exception as e:
        print(f"  [오류] 페이지 로드 실패: {e}")
        return "fail_load"
    page.wait_for_timeout(4000)
    title = post.get("title", "")
    body_snip = read_post_snippet(page)
    print(f"  체류 {STAY_SEC}초…")
    page.wait_for_timeout(STAY_SEC * 1000)
    comment = pick_comment(title, body_snip, comment_pool)
    liked = like_post(page)
    commented = write_comment(page, comment, post_url=post["url"], title=title)
    _append_jsonl(VISIT_LOG, {
        "date": date.today().isoformat(),
        "blog_id": blog_id,
        "group": group,
        "post_url": post["url"],
        "post_title": title,
        "comment": comment if commented else "",
        "liked": liked,
        "commented": commented,
        "action": "visit",
    })
    visited.add(blog_id)
    update_buddy_exchange(blog_id, visited_me=False)
    return "ok" if (commented and liked) else "partial"


def update_buddy_exchange(blog_id: str, visited_me: bool = False) -> None:
    if not BUDDY_XLSX.exists():
        return
    wb = openpyxl.load_workbook(BUDDY_XLSX)
    ws = wb.active
    today = date.today().isoformat()
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, 1).value or "").strip() == blog_id:
            if visited_me:
                ws.cell(r, 5, today)
            else:
                ws.cell(r, 4, today)
            cnt = int(ws.cell(r, 6).value or 0) + 1
            ws.cell(r, 6, cnt)
            wb.save(BUDDY_XLSX)
            return


def load_replied_today() -> set[str]:
    today = date.today().isoformat()
    seen: set[str] = set()
    if not REPLY_LOG.exists():
        return seen
    for line in REPLY_LOG.read_text(encoding="utf-8").splitlines():
        try:
            e = json.loads(line)
        except json.JSONDecodeError:
            continue
        if e.get("date") == today and e.get("commenter_id"):
            seen.add(e["commenter_id"])
    return seen


def collect_my_recent_post_urls(page: Page, limit: int = 5) -> list[dict]:
    page.goto(
        f"https://blog.naver.com/PostList.naver?blogId={MY_BLOG_ID}&categoryNo=0",
        wait_until="domcontentloaded",
        timeout=25000,
    )
    page.wait_for_timeout(3000)

    def _scan(root) -> list[dict]:
        return root.evaluate(
            """(params) => {
            const bid = params.bid;
            const lim = params.lim;
            const out = [];
            const seen = new Set();
            document.querySelectorAll('a[href*="'+bid+'/"]').forEach(a => {
              const h = (a.href || '').split('#')[0];
              if (!/blog\\.naver\\.com\\/[^/]+\\/\\d+$/.test(h) || seen.has(h)) return;
              seen.add(h);
              const m = h.match(/\\/(\\d+)$/);
              out.push({ url: h, log_no: m ? m[1] : '', title: (a.innerText||'').trim().slice(0,80) });
            });
            return out.slice(0, lim);
        }""",
            {"bid": MY_BLOG_ID, "lim": limit},
        )

    urls = _scan(page)
    if not urls:
        for fr in page.frames:
            if fr.url == "about:blank":
                continue
            try:
                urls = _scan(fr)
                if urls:
                    break
            except Exception:
                pass
    if urls:
        return urls
    try:
        fl = _main_frame(page)
        href = fl.locator("body").evaluate(
            """(bid) => {
            const links = [...document.querySelectorAll('a[href*="'+bid+'/"]')];
            for (const a of links) {
              const h = a.href.split('#')[0];
              if (/\\/\\d+$/.test(h)) return [{ url: h, log_no: h.split('/').pop(), title: '' }];
            }
            return [];
        }""",
            MY_BLOG_ID,
        )
        return href or []
    except Exception:
        return []


def collect_commenters_on_post(page: Page, log_no: str) -> list[dict]:
    from naver_comment import mobile_comment_url

    url = mobile_comment_url(MY_BLOG_ID, log_no)
    page.goto(url, wait_until="domcontentloaded", timeout=25000)
    page.wait_for_timeout(2500)
    rows = page.evaluate(
        """(myId) => {
        const out = [];
        const seen = new Set();
        document.querySelectorAll('.u_cbox_comment, .u_cbox_comment_box').forEach(box => {
          const nickEl = box.querySelector('.u_cbox_nick a, .u_cbox_nick');
          const nick = (nickEl?.innerText || '').trim();
          if (!nick) return;
          let bid = '';
          const href = nickEl?.href || '';
          let m = href.match(/blogId=([^&]+)/);
          if (m) bid = m[1];
          m = href.match(/blog\\.naver\\.com\\/([^/?#]+)/);
          if (!bid && m) bid = m[1];
          if (!bid || bid === myId || seen.has(bid)) return;
          seen.add(bid);
          out.push({ nick, blog_id: bid });
        });
        return out;
    }""",
        MY_BLOG_ID,
    )
    return rows or []


def reply_on_my_post(page: Page, log_no: str, nick: str, text: str) -> bool:
    from naver_comment import mobile_comment_url

    url = mobile_comment_url(MY_BLOG_ID, log_no)
    page.goto(url, wait_until="domcontentloaded", timeout=25000)
    page.wait_for_timeout(2000)
    ok = page.evaluate(
        """(nick, msg) => {
        const boxes = [...document.querySelectorAll('.u_cbox_comment, .u_cbox_comment_box')];
        for (const box of boxes) {
          const n = (box.querySelector('.u_cbox_nick a, .u_cbox_nick')?.innerText || '').trim();
          if (n !== nick) continue;
          const replyBtn = box.querySelector('.u_cbox_btn_reply, button.u_cbox_btn_reply, a.u_cbox_btn_reply');
          if (replyBtn) replyBtn.click();
          const el = document.querySelector('#naverComment__write_textarea, .u_cbox_text[contenteditable="true"]');
          if (!el) return false;
          el.focus();
          el.textContent = msg;
          el.dispatchEvent(new Event('input', { bubbles: true }));
          const btn = document.querySelector('.u_cbox_btn_upload, button.u_cbox_btn_upload');
          if (!btn) return false;
          btn.click();
          return true;
        }
        return false;
    }""",
        nick,
        text,
    )
    page.wait_for_timeout(1500)
    return bool(ok)


def run_reply_commenters(page: Page, max_commenters: int = 15) -> None:
    """내 글에 댓글 남긴 이웃 → 답글 + 그 이웃 최신글 방문(댓글+하트)."""
    ensure_login(page)
    comments = load_comments()
    replied = load_replied_today()
    posts = collect_my_recent_post_urls(page, limit=5)
    print(f"[답글] 내 최근글 {len(posts)}건 스캔")
    targets: list[dict] = []
    seen: set[str] = set()
    for post in posts:
        log_no = post.get("log_no", "")
        if not log_no:
            continue
        for row in collect_commenters_on_post(page, log_no):
            bid = row.get("blog_id", "")
            if not bid or bid in seen or bid in replied or bid in SKIP_BLOG_IDS:
                continue
            seen.add(bid)
            targets.append({
                "blog_id": bid,
                "nick": row.get("nick", ""),
                "log_no": log_no,
                "post_url": post.get("url", ""),
            })
            if len(targets) >= max_commenters:
                break
        if len(targets) >= max_commenters:
            break

    print(f"[답글] 대상 {len(targets)}명")
    done = 0
    for t in targets:
        bid = t["blog_id"]
        nick = t["nick"]
        reply_text = random.choice(REPLY_MESSAGES)
        print(f"[답글] {bid} ({nick})")
        replied_ok = reply_on_my_post(page, t["log_no"], nick, reply_text)
        print(f"  내글 답글: {'성공' if replied_ok else '실패'}")

        neighbor_post = get_latest_post(page, bid)
        visited_ok = liked = commented = False
        if neighbor_post:
            page.goto(neighbor_post["url"], wait_until="domcontentloaded", timeout=25000)
            page.wait_for_timeout(3000)
            body = read_post_snippet(page)
            comment = pick_comment(neighbor_post.get("title", ""), body, comments)
            liked = like_post(page)
            commented = write_comment(
                page, comment, post_url=neighbor_post["url"], title=neighbor_post.get("title", "")
            )
            visited_ok = liked or commented
            print(f"  이웃글 방문: like={liked} comment={commented}")

        _append_jsonl(REPLY_LOG, {
            "date": date.today().isoformat(),
            "commenter_id": bid,
            "nick": nick,
            "my_log_no": t["log_no"],
            "reply_text": reply_text if replied_ok else "",
            "replied": replied_ok,
            "visited_neighbor": visited_ok,
            "liked": liked,
            "commented": commented,
            "neighbor_url": neighbor_post["url"] if neighbor_post else "",
        })
        if replied_ok or visited_ok:
            done += 1
            update_buddy_exchange(bid, visited_me=True)
        time.sleep(random.uniform(2, 4))
    print(f"[답글 완료] {done}/{len(targets)}명 처리")


def run_retry_partial(page: Page, all_dates: bool = True) -> None:
    """partial 방문 건 재시도 — 댓글/하트 미완료 건."""
    if not VISIT_LOG.exists():
        return
    comments = load_comments()
    today = date.today().isoformat()
    seen_urls: set[str] = set()
    targets = []
    for line in VISIT_LOG.read_text(encoding="utf-8").splitlines():
        try:
            e = json.loads(line)
        except json.JSONDecodeError:
            continue
        if e.get("action") != "visit":
            continue
        if not all_dates and e.get("date") != today:
            continue
        url = e.get("post_url", "")
        if url in seen_urls:
            continue
        if e.get("commented") and e.get("liked"):
            continue
        seen_urls.add(url)
        targets.append(e)
    print(f"[재시도] partial {len(targets)}건 (all_dates={all_dates})")
    ok_c = ok_l = 0
    for e in targets:
        post = {"url": e["post_url"], "title": e.get("post_title", ""), "blog_id": e["blog_id"]}
        print(f"  {e['blog_id']}")
        if not e.get("liked"):
            try:
                page.goto(post["url"], wait_until="domcontentloaded", timeout=25000)
                page.wait_for_timeout(3000)
            except Exception as ex:
                print(f"    [오류] {ex}")
                continue
        comment = pick_comment(post["title"], "", comments) if not e.get("commented") else ""
        if not e.get("commented") and not comment:
            body = ""
            try:
                page.goto(post["url"], wait_until="domcontentloaded", timeout=25000)
                page.wait_for_timeout(3000)
                body = read_post_snippet(page)
            except Exception:
                pass
            comment = pick_comment(post["title"], body, comments)
        liked = like_post(page) if not e.get("liked") else True
        commented = (
            write_comment(page, comment, post_url=post["url"], title=post["title"])
            if not e.get("commented") else True
        )
        if liked:
            ok_l += 1
        if commented:
            ok_c += 1
        _append_jsonl(VISIT_LOG, {
            "date": date.today().isoformat(),
            "blog_id": e["blog_id"],
            "post_url": e["post_url"],
            "post_title": e.get("post_title", ""),
            "comment": comment if commented else "",
            "liked": liked,
            "commented": commented,
            "action": "retry",
        })
        print(f"    like={liked} comment={commented}")
        time.sleep(random.uniform(2, 4))
    print(f"[재시도 완료] 하트 {ok_l}/{len(targets)} | 댓글 {ok_c}/{len(targets)}")


def collect_all_it_posts(page: Page, max_pages: int = 30) -> list[dict]:
    """IT 그룹 이웃새글 전 페이지 순회 → blog_id당 최신글 1개."""
    select_neighbor_group(page, "it")
    by_id: dict[str, dict] = {}
    for pg_num in range(1, max_pages + 1):
        url = (
            f"https://section.blog.naver.com/connect/ViewMoreBuddyPosts.naver"
            f"?blogId={MY_BLOG_ID}&currentPage={pg_num}"
        )
        page.goto(url, wait_until="domcontentloaded", timeout=25000)
        page.wait_for_timeout(1500)
        # IT 그룹 재선택 (페이지 이동 시 초기화될 수 있음)
        label = GROUP_LABELS["it"]
        sel = page.locator("select").first
        if sel.count():
            try:
                sel.select_option(label=label)
                page.wait_for_timeout(1500)
            except Exception:
                pass
        batch = collect_post_links(page, limit=200)
        if not batch:
            print(f"[수집] p{pg_num} — 글 없음, 종료")
            break
        new = 0
        for post in batch:
            bid = post["blog_id"]
            if bid not in by_id:
                by_id[bid] = post
                new += 1
        print(f"[수집] p{pg_num} +{new}명 (누적 {len(by_id)}명)")
        if new == 0:
            break
    return list(by_id.values())


def get_latest_post(page: Page, blog_id: str) -> dict | None:
    """블로그 최신글 URL 탐색."""
    page.goto(f"https://blog.naver.com/{blog_id}", wait_until="domcontentloaded", timeout=25000)
    page.wait_for_timeout(3000)
    try:
        fl = _main_frame(page)
        href = fl.locator("body").evaluate("""() => {
          const a = document.querySelector('a[href*="/'+location.pathname.split('/').pop()+'/"]')
            || document.querySelector('a.link_title, .blog2_series a, a[href*="logNo="]');
          if (a) return a.href.split('#')[0];
          const links = [...document.querySelectorAll('a[href*="blog.naver.com"]')];
          for (const l of links) {
            if (/\\/\\d+$/.test(l.href.split('#')[0])) return l.href.split('#')[0];
          }
          return '';
        }""")
        if href and blog_id in href:
            return {"url": href, "blog_id": blog_id, "title": f"{blog_id} 최신글"}
    except Exception:
        pass
    posts = page.evaluate(
        """(bid) => {
        const out = [];
        document.querySelectorAll('a[href*="blog.naver.com/'+bid+'/"]').forEach(a => {
          const h = a.href.split('#')[0];
          if (/\\/\\d+$/.test(h)) out.push(h);
        });
        return out[0] || '';
    }""",
        blog_id,
    )
    if posts:
        return {"url": posts, "blog_id": blog_id, "title": f"{blog_id} 최신글"}
    return None


def run_visit_it_all(page: Page) -> None:
    """IT 그룹 이웃 전원 — 블로그당 1회, 댓글+하트."""
    ensure_login(page)
    comments = load_comments()
    visited = load_visited_today()
    posts = collect_all_it_posts(page)
    print(f"[IT 전체] 대상 {len(posts)}명 / 오늘 이미 방문 {len(visited)}명")
    done = ok = 0
    for post in posts:
        bid = post["blog_id"]
        if bid in visited:
            print(f"[skip] {bid} — 오늘 방문함")
            continue
        print(f"[방문] {bid} — {post.get('title', '')[:45]}")
        result = visit_one_post(page, post, comments, visited, GROUP_LABELS["it"])
        print(f"  → {result}")
        done += 1
        if result == "ok":
            ok += 1
        time.sleep(random.uniform(2, 4))
    print(f"[완료] IT 전체 방문 {done}건 (완전성공 {ok}건)")


def run_visit_group(page: Page, group_key: str, limit: int) -> None:
    ensure_login(page)
    comments = load_comments()
    visited = load_visited_today()
    select_neighbor_group(page, group_key)
    posts = collect_post_links(page, limit=limit * 5)
    done = 0
    for post in posts:
        if done >= limit:
            break
        if post["blog_id"] in visited:
            print(f"[skip] 오늘 방문함: {post['blog_id']}")
            continue
        print(f"[방문] {post['blog_id']} — {post['title'][:45]}")
        result = visit_one_post(page, post, comments, visited, GROUP_LABELS.get(group_key, group_key))
        print(f"  → {result}")
        if result in ("ok", "partial"):
            done += 1
        time.sleep(random.uniform(2, 4))
    print(f"[완료] {GROUP_LABELS.get(group_key)} 방문 {done}건")


def parse_neighbor_count(page: Page) -> int | None:
    try:
        txt = page.locator("body").inner_text(timeout=5000)
        m = re.search(r"이웃\s*([\d,]+)", txt)
        if m:
            return int(m.group(1).replace(",", ""))
    except Exception:
        pass
    for fr in page.frames:
        try:
            txt = fr.locator("body").inner_text(timeout=2000)
            m = re.search(r"이웃\s*([\d,]+)", txt)
            if m:
                return int(m.group(1).replace(",", ""))
        except Exception:
            pass
    return None


def collect_neighbor_candidates(page: Page, per_theme: int = 5) -> list[str]:
    found: list[str] = []
    seen: set[str] = load_applied_ids() | {MY_BLOG_ID}
    invalid = SKIP_BLOG_IDS | {
        "BlogHome.naver", "MyBlog.naver", "market", "blogpeople", "PostView.naver",
    }
    for name, url in THEME_DIRS.items():
        page.goto(url, wait_until="domcontentloaded", timeout=25000)
        page.wait_for_timeout(2000)
        ids = page.evaluate(
            """() => {
            const out=[]; const s=new Set();
            document.querySelectorAll('a[href*="blog.naver.com/"]').forEach(a=>{
              const m=a.href.match(/blog\\.naver\\.com\\/([^/?#]+)\\/(\\d+)/);
              if(!m||s.has(m[1])) return;
              s.add(m[1]); out.push(m[1]);
            });
            return out;
        }"""
        )
        added = 0
        for bid in ids:
            if bid in seen or bid in invalid or len(bid) < 3:
                continue
            if not re.match(r"^[a-zA-Z0-9_-]+$", bid):
                continue
            found.append(bid)
            seen.add(bid)
            added += 1
            if added >= per_theme:
                break
        print(f"[후보] {name} 테마 +{added}명")
    return found


def apply_neighbor(page: Page, blog_id: str) -> dict:
    url = f"https://blog.naver.com/{blog_id}"
    page.goto(url, wait_until="domcontentloaded", timeout=25000)
    page.wait_for_timeout(2000)
    count = parse_neighbor_count(page)
    if count is not None and count >= MAX_NEIGHBORS:
        return {"blog_id": blog_id, "status": "skip_full", "neighbor_count": count}

    root = _main_frame(page)
    for sel in (
        "a:has-text('이웃추가')",
        "button:has-text('이웃추가')",
        "a._buddy_add",
        "a.btn_buddy",
    ):
        loc = root.locator(sel)
        if loc.count():
            try:
                loc.first.click(timeout=3000)
                page.wait_for_timeout(1500)
                for confirm in ("a:has-text('서로이웃')", "button:has-text('서로이웃')", "a:has-text('이웃신청')", "button:has-text('확인')"):
                    c = page.locator(confirm)
                    if c.count():
                        c.first.click(timeout=2000)
                        page.wait_for_timeout(1000)
                        break
                for msg_sel in ("textarea[placeholder*='인사']", "textarea"):
                    ta = page.locator(msg_sel)
                    if ta.count() and ta.first.is_visible():
                        ta.first.fill("안녕하세요~ IT·가전 글 잘 보고 있습니다. 이웃 신청드려요 😊")
                        page.locator("button:has-text('확인'), a:has-text('확인')").first.click(timeout=2000)
                        page.wait_for_timeout(1000)
                        break
                entry = {"date": date.today().isoformat(), "blog_id": blog_id, "status": "applied", "neighbor_count": count}
                _append_jsonl(APPLY_LOG, entry)
                return entry
            except Exception as e:
                return {"blog_id": blog_id, "status": "fail", "error": str(e)[:80]}
    return {"blog_id": blog_id, "status": "skip_exists_or_ui", "neighbor_count": count}


def run_apply_neighbors(page: Page, limit: int) -> None:
    ensure_login(page)
    candidates = collect_neighbor_candidates(page, per_theme=8)
    done = 0
    for blog_id in candidates:
        if done >= limit:
            break
        print(f"[이웃신청] {blog_id}")
        result = apply_neighbor(page, blog_id)
        print(f"  → {result['status']}")
        if result["status"] == "applied":
            done += 1
            add_buddy_row(blog_id, "그냥이웃")
        time.sleep(random.uniform(3, 6))
    print(f"[완료] 이웃신청 {done}/{limit}건")


def add_buddy_row(blog_id: str, group: str) -> None:
    if not BUDDY_XLSX.exists():
        return
    wb = openpyxl.load_workbook(BUDDY_XLSX)
    ws = wb.active
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, 1).value or "").strip() == blog_id:
            return
    ws.append([blog_id, "", group, "", "", 0, "", "자동신청"])
    wb.save(BUDDY_XLSX)


def open_write_editor(page: Page) -> None:
    page.goto(f"https://blog.naver.com/{MY_BLOG_ID}?Redirect=Write&", wait_until="domcontentloaded", timeout=30000)
    page.wait_for_timeout(6000)


def _write_frame(page: Page):
    for fr in page.frames:
        if "PostWriteForm" in fr.url:
            return fr
    return page


def run_save_draft(page: Page, html_path: Path, image_names: list[str]) -> bool:
    ensure_login(page)
    if not html_path.exists():
        raise FileNotFoundError(html_path)
    html = html_path.read_text(encoding="utf-8")
    title_m = re.search(r"<title>([^<]+)</title>", html, re.I)
    title = title_m.group(1).strip() if title_m else "제목 없음"
    body_m = re.search(r"<body[^>]*>(.*)</body>", html, re.I | re.S)
    body_html = body_m.group(1) if body_m else html
    plain = re.sub(r"<br\s*/?>", "\n", body_html, flags=re.I)
    plain = re.sub(r"</p>", "\n\n", plain, flags=re.I)
    plain = re.sub(r"<[^>]+>", "", plain)
    plain = re.sub(r"\n{3,}", "\n\n", plain).strip()

    open_write_editor(page)
    page.wait_for_timeout(8000)
    fr = _write_frame(page)

    # 제목
    for sel in (".se-title-text", ".se-section-documentTitle [contenteditable='true']"):
        loc = fr.locator(sel)
        if loc.count():
            try:
                loc.first.click(timeout=5000)
                page.keyboard.press("Control+A")
                page.keyboard.insert_text(title)
                break
            except Exception:
                try:
                    loc.first.fill(title)
                    break
                except Exception:
                    pass

    # 본문 — documentTitle 다음 text 섹션
    body_filled = False
    for sel in (
        ".se-section-text .se-text-paragraph",
        ".se-component.se-text .se-text-paragraph",
    ):
        loc = fr.locator(sel)
        if loc.count():
            try:
                loc.first.click(timeout=5000)
                page.keyboard.insert_text(plain[:8000])
                body_filled = True
                break
            except Exception:
                pass
    if not body_filled:
        loc = fr.locator(".se-text-paragraph")
        if loc.count() >= 2:
            try:
                loc.nth(1).click(timeout=5000)
                page.keyboard.insert_text(plain[:8000])
                body_filled = True
            except Exception:
                pass
    if not body_filled:
        for sel in (".se-section-text [contenteditable='true']", "[contenteditable='true']"):
            loc = fr.locator(sel)
            if loc.count():
                try:
                    loc.last.click(timeout=5000)
                    page.keyboard.insert_text(plain[:8000])
                    body_filled = True
                    break
                except Exception:
                    pass

    page.wait_for_timeout(1500)

    # 이미지 업로드
    for img_name in image_names:
        img_path = IMAGES / img_name
        if not img_path.exists():
            img_path = ROOT / img_name
        if not img_path.exists():
            continue
        for sel in (
            "button.se-image-toolbar-button",
            "button:has-text('사진')",
            "[data-name='image']",
        ):
            btn = fr.locator(sel)
            if btn.count():
                try:
                    with page.expect_file_chooser(timeout=5000) as fc:
                        btn.first.click()
                    fc.value.set_files(str(img_path))
                    page.wait_for_timeout(2500)
                except Exception:
                    pass
                break

    # 임시저장 — Smart Editor ONE: 상단 '저장' 버튼 (save_btn__*)
    save_selectors = (
        "button.save_btn__bzc5B",
        "button[class*='save_btn__']",
        ".save_btn_area__Qo0W7 button",
        "button:has-text('임시저장')",
        "span:has-text('임시저장')",
        ".save_count_btn",
        "button[data-click-area='tpb.save']",
    )
    for root in (fr, page):
        for sel in save_selectors:
            btn = root.locator(sel)
            if not btn.count():
                continue
            try:
                btn.first.click(timeout=5000)
                page.wait_for_timeout(3000)
                _append_jsonl(LOGS / "draft_log.jsonl", {
                    "date": date.today().isoformat(),
                    "title": title,
                    "html": str(html_path),
                    "status": "draft_clicked",
                    "body_filled": body_filled,
                })
                print(f"[임시저장] '{title}' 클릭 완료")
                return True
            except Exception:
                pass

    print("[임시저장] 버튼 못 찾음 — 수동 확인 필요")
    try:
        fr.screenshot(path=str(LOGS / "draft_fail.png"), timeout=15000)
    except Exception:
        pass
    return False


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--mode", choices=["visit-it", "visit-it-all", "apply", "draft", "probe", "retry-partial", "test-comment", "reply-commenters"], default="probe")
    ap.add_argument("--url", type=str, default="", help="test-comment용 글 URL")
    ap.add_argument("--message", type=str, default="", help="test-comment용 댓글 문장")
    ap.add_argument("--limit", type=int, default=3)
    ap.add_argument("--html", type=str, default="")
    ap.add_argument("--images", type=str, default="", help="쉼표 구분 이미지 파일명")
    args = ap.parse_args()

    with sync_playwright() as p:
        browser = p.chromium.connect_over_cdp(CDP)
        page = get_cdp_page(browser)
        if args.mode == "visit-it":
            run_visit_group(page, "it", args.limit)
        elif args.mode == "visit-it-all":
            run_visit_it_all(page)
        elif args.mode == "apply":
            run_apply_neighbors(page, args.limit)
        elif args.mode == "retry-partial":
            ensure_login(page)
            run_retry_partial(page)
        elif args.mode == "test-comment":
            ensure_login(page)
            url = args.url or "https://blog.naver.com/gncomblog/224326086074"
            msg = args.message or "와~꿀팁이네요 저장해뒀어요 ㅎㅎ 좋은하루 보내세요"
            ok = write_neighbor_comment(page, url, msg, log_path=COMMENT_LOG)
            print(f"[test-comment] {'성공' if ok else '실패'}")
        elif args.mode == "reply-commenters":
            run_reply_commenters(page, max_commenters=args.limit)
        elif args.mode == "draft":
            html = Path(args.html) if args.html else sorted(POSTS.glob("*.html"))[-1]
            imgs = [x.strip() for x in args.images.split(",") if x.strip()] if args.images else []
            run_save_draft(page, html, imgs)
        else:
            ensure_login(page)
            select_neighbor_group(page, "it")
            posts = collect_post_links(page, 5)
            print(f"[probe] {len(posts)}건")
            for po in posts:
                print(f"  {po['blog_id']}: {po['title'][:50]}")


if __name__ == "__main__":
    main()
