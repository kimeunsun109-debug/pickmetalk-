#!/usr/bin/env python3
"""이웃 분류·교류 관리 엑셀 생성"""
from pathlib import Path
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

OUT = Path(__file__).resolve().parent.parent / "data" / "이웃분류_교류관리.xlsx"
HEADERS = [
    "blog_id", "블로그명", "그룹", "마지막_내방문", "마지막_상대방문",
    "교류횟수", "이웃수", "비고",
]
GROUPS = "IT 관련 | 재테크 | 일상글 | 그냥이웃"


def main():
    if OUT.exists():
        print(f"이미 존재: {OUT}")
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "이웃분류"
    fill = PatternFill("solid", fgColor="E8837A")
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
    ws.cell(2, 8, f"그룹: {GROUPS}")
    ws.cell(3, 8, "그냥이웃=먼저방문X, 내글 반응시만 답방")
    ws.cell(4, 8, "IT/재테크/일상=티키타카 교류, 한달 무응답→그냥이웃")
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 8
    ws.column_dimensions["H"].width = 36
    wb.save(OUT)
    print(f"생성: {OUT}")


if __name__ == "__main__":
    main()
