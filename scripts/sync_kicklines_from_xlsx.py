from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
XLSX_PATH = ROOT / "킥문장_마스터DB.xlsx"
OUT_PATH = ROOT / "data" / "kickLines" / "master.json"

TYPE_TO_CATEGORY = {
    "감동킥": "touching",
    "유머킥": "joke",
    "팩트킥": "wit",
    "드립킥": "loving_nag",
    "감성킥": "flutter",
    "관찰킥": "comfort",
}

DEFAULT_LABELS = {
    "touching": "감동 한방",
    "comfort": "위로",
    "flutter": "설렘",
    "joke": "유머",
    "loving_nag": "사랑 담긴 잔소리",
    "wit": "센스/위트",
    "closing": "마무리",
}


def normalize(v: object) -> str:
    return str(v or "").strip()


def main() -> None:
    if not XLSX_PATH.exists():
        if OUT_PATH.exists():
            print(f"skip: {XLSX_PATH.name} not found, using existing {OUT_PATH.name}")
            return
        raise FileNotFoundError(f"missing workbook: {XLSX_PATH}")

    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb["⚡ 전체 킥문장 마스터"]

    by_category: dict[str, list[str]] = {k: [] for k in DEFAULT_LABELS.keys()}
    seen: dict[str, set[str]] = {k: set() for k in DEFAULT_LABELS.keys()}

    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row:
            continue
        kick_type = normalize(row[3] if len(row) > 3 else "")
        line = normalize(row[4] if len(row) > 4 else "")
        if not line:
            continue

        category = TYPE_TO_CATEGORY.get(kick_type)
        if not category:
            continue

        if line not in seen[category]:
            seen[category].add(line)
            by_category[category].append(line)

    wb.close()

    packs = []
    for category, label in DEFAULT_LABELS.items():
        lines = by_category.get(category, [])
        if not lines:
            continue
        packs.append({"category": category, "label": label, "lines": lines})

    payload = {
        "source": XLSX_PATH.name,
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "packs": packs,
    }
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    print(f"saved: {OUT_PATH}")
    print(f"packs: {len(packs)}")
    for p in packs:
        print(f"  {p['category']}: {len(p['lines'])}")


if __name__ == "__main__":
    main()
