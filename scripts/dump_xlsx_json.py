import json
from pathlib import Path
from openpyxl import load_workbook

p = Path("character_dialogue_examples.xlsx")
wb = load_workbook(p, read_only=True)
out = {"sheets": []}
for name in wb.sheetnames:
    ws = wb[name]
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([str(c) if c is not None else None for c in row])
    out["sheets"].append({"name": name, "rows": len(rows), "sample": rows[:12]})
wb.close()
Path("scripts/xlsx_dump.json").write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
print("done", len(out["sheets"]))
