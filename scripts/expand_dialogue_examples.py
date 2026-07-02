"""
character_dialogue_examples.xlsx 확장
- 사용자 샘플 유지
- 카테고리별 100개 대화 pair (캐릭터 시트)
- 대화패턴 컬럼 추가
"""

from __future__ import annotations

import copy
import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
XLSX_PATH = ROOT / "character_dialogue_examples.xlsx"

CHARACTER_SHEETS: dict[str, str] = {
    "💛 유나": "yuna",
    "🖤 나린": "narin",
    "📋 윤서": "yoonseo",
    "🌙 은하": "eunha",
    "⚡ 지유": "jiyu",
}

CHARACTER_COLS = ["💛 유나", "🖤 나린", "📋 윤서", "🌙 은하", "⚡ 지유"]

CATEGORY_TOPICS: dict[str, list[str]] = {
    "감정 상태": [
        "짜증남",
        "무기력함",
        "외로움",
        "설렘",
        "뿌듯함",
        "불안함",
        "슬픔",
        "멍함",
    ],
    "일상 이벤트": [
        "야근",
        "퇴근",
        "점심",
        "주말",
        "월요일 아침",
        "잠 못 잠",
        "운동함",
        "쉬는 날",
        "불금",
    ],
    "관계/사람": [
        "친구에게 상처받음",
        "누군가 보고 싶음",
        "싸움",
        "칭찬받음",
        "무시당한 느낌",
        "반려동물의 죽음",
    ],
    "시간대/날씨": [
        "새벽 감성",
        "비 오는 날",
        "첫눈",
        "더운 여름 낮",
        "명절",
        "대체공휴일",
    ],
    "자기 자신": [
        "자신감 없음",
        "뭔가 이루고 싶음",
        "나태해진 느낌",
        "잘 해냈을 때",
        "머리가 복잡함",
    ],
    "대화 행동": [
        "사과",
        "달래기",
        "축하",
        "칭찬",
        "주제 전환",
        "질문형 마무리",
        "여운형 마무리",
        "제안형 마무리",
    ],
}

CATEGORY_SHEET_NAMES: dict[str, str] = {
    "감정 상태": "📁 감정 상태",
    "일상 이벤트": "📁 일상 이벤트",
    "관계/사람": "📁 관계·사람",
    "시간대/날씨": "📁 시간대·날씨",
    "자기 자신": "📁 자기 자신",
    "대화 행동": "📁 대화 행동",
}

DIALOGUE_PATTERNS = [
    "공감",
    "배려",
    "유머",
    "관찰",
    "질문",
    "제안",
    "드립",
    "여운",
    "내적독백",
    "정보",
]

USER_UTTERANCES: dict[str, list[str]] = {
    "짜증남": [
        "아 진짜 오늘 너무 짜증나",
        "하... 열받아서 말이 안 나와",
        "오늘 빌런 제대로 만났어",
        "말투가 심상치 않은데",
        "진짜 어이없는 일 있었어",
        "화가 치밀어 올라",
        "오늘은 왜 이렇게 꼬이지",
        "답답해서 미치겠다",
    ],
    "무기력함": [
        "요즘 아무것도 하기 싫어",
        "에너지가 바닥이야",
        "몸도 마음도 축 처진다",
        "의욕이 0이야",
        "그냥 누워있고 싶다",
        "아무것도 손에 안 잡혀",
    ],
    "외로움": [
        "오늘따라 너무 외롭다",
        "괜히 마음이 비어있는 느낌이야",
        "사람 많은데도 혼자인 기분이야",
        "연락할 사람이 없는 느낌",
        "갑자기 쓸쓸해졌어",
    ],
    "설렘": [
        "오늘 좀 설레는 일이 생겼어",
        "괜히 기분이 들뜨네",
        "심장이 빨리 뛰는 느낌이야",
        "좋은 소식 들었어",
        "왠지 기대되는 하루야",
    ],
    "뿌듯함": [
        "오늘 진짜 잘한 것 같아",
        "드디어 해냈다",
        "내가 좀 대견하다",
        "끝내고 나니 뿌듯하네",
        "오늘은 나 칭찬해줘도 돼",
    ],
    "불안함": [
        "괜히 불안해서 집중이 안 돼",
        "머릿속이 복잡해",
        "아직 안 일어난 일인데 걱정돼",
        "마음이 조마조마해",
        "불안이 자꾸 올라와",
    ],
    "슬픔": [
        "오늘 좀 많이 슬프네",
        "눈물 날 것 같다",
        "마음이 너무 가라앉아",
        "울컥하네",
        "그냥 슬퍼",
    ],
    "멍함": [
        "멍하게 시간만 흘렀어",
        "오늘 정신이 하나도 없어",
        "뭘 했는지 기억이 안 나",
        "멍 때리고 있었어",
        "뇌가 멈춘 느낌",
    ],
    "야근": [
        "오늘 또 야근 확정이야",
        "퇴근이 안 보인다",
        "지금도 회사야",
        "아...진짜...ㅜㅜ 오늘 또 야근이야..",
        "밤 10시 넘었는데 아직 일해",
    ],
    "퇴근": [
        "이제 퇴근한다",
        "오늘도 겨우 버텼네",
        "집 가는 길이야",
        "드디어 나왔다",
        "퇴근길이라 기분 좋아",
    ],
    "점심": [
        "점심 뭐 먹지",
        "배는 고픈데 귀찮다",
        "오늘 점심 추천해줘",
        "밥 먹을 시간인데 메뉴 고민",
        "점심시간만 기다렸어",
    ],
    "주말": [
        "주말인데 뭘 해야 할지 모르겠어",
        "이번 주말은 쉬고 싶다",
        "주말 계획 아직 없음",
        "주말인데 심심해",
        "이번 주말 너무 바빴어",
    ],
    "월요일 아침": [
        "월요일 아침부터 힘들다",
        "출근길부터 텐션 바닥",
        "월요병 제대로 왔어",
        "월요일이라 벌써 지쳐",
        "출근하기 싫다",
    ],
    "잠 못 잠": [
        "어제 잠을 거의 못 잤어",
        "새벽까지 뒤척였다",
        "눈은 감았는데 머리가 안 쉬더라",
        "불면이라 피곤해",
        "잠이 안 와",
    ],
    "운동함": [
        "오늘 운동하고 왔어",
        "땀 좀 빼니까 살 것 같다",
        "오랜만에 몸 풀었다",
        "운동했더니 개운해",
        "헬스 갔다 왔어",
    ],
    "쉬는 날": [
        "오늘은 쉬는 날이야",
        "아무 일정 없이 쉬고 있다",
        "오늘은 집콕 모드",
        "푹 쉬는 날",
        "일 안 하는 날이 최고",
    ],
    "불금": [
        "불금인데 뭐하지",
        "오늘은 좀 놀고 싶다",
        "불금 텐션 올려줘",
        "불금이라 설레",
        "금요일 밤이다",
    ],
    "친구에게 상처받음": [
        "친구한테 말로 상처받았어",
        "가까운 사람한테 실망했다",
        "친구 때문에 마음이 아프다",
        "친한 친구한테 서운했어",
        "친구 말이 너무 셌어",
    ],
    "누군가 보고 싶음": [
        "갑자기 누가 보고 싶다",
        "문득 생각나는 사람이 있어",
        "그 사람이 자꾸 떠오른다",
        "보고 싶은 사람이 생겼어",
        "연락하고 싶은데 참았어",
    ],
    "싸움": [
        "아까 크게 다퉜어",
        "말하다가 싸워버렸어",
        "감정이 너무 올라왔어",
        "연인한테 싸웠어",
        "말다툼했어",
    ],
    "칭찬받음": [
        "오늘 칭찬받았어",
        "생각보다 좋은 평가 받았다",
        "인정받으니까 기분 좋다",
        "칭찬 들었어",
        "오늘 인정받은 느낌",
    ],
    "무시당한 느낌": [
        "오늘 무시당한 기분 들었어",
        "내 말이 투명해진 느낌이야",
        "존중 못 받은 느낌이라 별로다",
        "말이 안 들리는 느낌",
        "내가 없는 사람 취급",
    ],
    "반려동물의 죽음": [
        "반려동물이 떠났어",
        "마음이 텅 빈 것 같다",
        "아직 실감이 안 난다",
        "우리 강아지가 없어졌어",
        "너무 슬퍼",
    ],
    "새벽 감성": [
        "새벽이라 그런지 생각이 많아",
        "이 시간엔 감정이 커진다",
        "잠도 안 오고 마음이 복잡해",
        "새벽 3시인데 눈 안 감겨",
        "밤이 길게 느껴져",
    ],
    "비 오는 날": [
        "비 오는 소리 들으니까 기분이 묘해",
        "오늘 하루 종일 비네",
        "비 오니까 괜히 울적해",
        "비 맞고 왔어",
        "우산 없이 나왔다",
    ],
    "첫눈": [
        "오늘 첫눈 봤어",
        "창밖에 눈 오는데 예쁘더라",
        "첫눈 보니까 마음이 몽글해",
        "눈 왔다",
        "첫눈이라 설렜어",
    ],
    "더운 여름 낮": [
        "너무 더워서 녹을 것 같다",
        "여름 낮 진짜 버티기 힘들다",
        "오늘 햇빛이 장난 아니야",
        "더위에 지쳤어",
        "에어컨 없으면 못 살아",
    ],
    "명절": [
        "명절이라 정신없다",
        "가족 모임이 좀 버겁다",
        "명절 분위기가 묘하게 복잡해",
        "명절 준비 지쳐",
        "친척 만나기 부담",
    ],
    "대체공휴일": [
        "대체공휴일인데 쉬는 게 어색하다",
        "하루 번 느낌이라 좋다",
        "오늘은 느긋하게 보내고 싶어",
        "쉬는 날인데 할 게 없네",
        "연휴 끝이라 아쉬워",
    ],
    "자신감 없음": [
        "요즘 자신감이 너무 떨어진다",
        "내가 잘하고 있는지 모르겠어",
        "괜히 다 작아 보인다",
        "자신감 바닥",
        "나한테 자신 없어",
    ],
    "뭔가 이루고 싶음": [
        "이번엔 진짜 뭐 하나 이루고 싶어",
        "변화가 필요해",
        "목표를 제대로 잡고 싶다",
        "새로 시작하고 싶어",
        "도전하고 싶은 게 있어",
    ],
    "나태해진 느낌": [
        "요즘 너무 늘어진 느낌이야",
        "해야 할 걸 자꾸 미룬다",
        "리듬이 완전 깨졌다",
        "게을러졌어",
        "미루기만 해",
    ],
    "잘 해냈을 때": [
        "이번 건 진짜 잘 해낸 것 같아",
        "끝내고 나니까 뿌듯하다",
        "고생한 보람이 있네",
        "오늘 잘했다",
        "목표 달성했어",
    ],
    "머리가 복잡함": [
        "머릿속이 너무 시끄러워",
        "정리해도 정리가 안 된다",
        "해야 할 게 많아서 멘붕이야",
        "생각이 너무 많아",
        "머리가 하얘져",
    ],
    "사과": [
        "아까 말 심했지 미안해",
        "내가 먼저 사과할게",
        "그건 내 실수였다",
        "미안, 좀 예민했어",
        "화낸 거 미안",
    ],
    "달래기": [
        "오늘 위로 좀 해줘",
        "따뜻한 말이 필요해",
        "마음이 너무 예민하다",
        "위로 받고 싶어",
        "힘들어 위로해줘",
    ],
    "축하": [
        "좋은 소식 있어",
        "이번에 진짜 잘됐다",
        "드디어 목표 달성했어",
        "축하할 일 생겼어",
        "기분 좋은 일 있었어",
    ],
    "칭찬": [
        "나 칭찬 좀 해줘",
        "오늘 나 괜찮지?",
        "잘한 거 맞지?",
        "나 좀 잘했어?",
        "칭찬 한마디 해줘",
    ],
    "주제 전환": [
        "분위기 바꿔볼까",
        "이 얘기는 여기까지 할래",
        "다른 얘기 하고 싶어",
        "주제 바꾸자",
        "가벼운 얘기 하자",
    ],
    "질문형 마무리": [
        "짧게 마무리하고 싶어",
        "한마디만 듣고 싶다",
        "오늘은 짧게 얘기하자",
        "한 줄만 더 말해줘",
        "마지막 한마디",
    ],
    "여운형 마무리": [
        "조용하게 끝내고 싶어",
        "오늘은 잔잔하게 마무리하고 싶다",
        "여운 남는 말 한 줄 해줘",
        "조용한 말 해줘",
        "잔잔하게 끝내자",
    ],
    "제안형 마무리": [
        "다음 액션 하나만 정해줘",
        "내일 할 거 한 개만 정하자",
        "작게라도 같이 해보자",
        "뭐부터 할지 정하자",
        "실행 하나만 골라줘",
    ],
}


def _pick(lines: list[str], variant: int) -> str:
    return lines[variant % len(lines)]


def render_reply(
    char_id: str, topic: str, pattern: str, variant: int, momentum: bool
) -> str:
    """캐릭터별 말투 + 대화패턴으로 응답 생성"""

    if char_id == "yuna":
        openers = {
            "공감": ["그 마음 이해해.", "말하는데 힘들었겠다.", "오늘 많이 버텼네."],
            "배려": ["일단 숨 고르자.", "밥은 먹었어?", "오늘은 네 편이야."],
            "유머": ["ㅋㅋ 그 말에 웃음 나온다.", "오늘 센스 있네.", "그거 재밌다."],
            "관찰": ["말투가 평소랑 달라.", "오늘 텐션이 다르다.", "눈치 보이는데."],
            "질문": ["그래서 어떻게 됐어?", "지금 제일 급한 건 뭐야?", "뭐가 그렇게 했어?"],
            "제안": ["같이 정리해볼까?", "밥부터 먹자.", "내가 들어줄게."],
            "드립": ["오늘은 살아남기 모드 ㅋ", "그건 좀 웃기네.", "헐 그건 반전이네."],
            "여운": ["오늘은 천천히 가자.", "괜찮아, 네 편이야.", "여기 있어."],
            "내적독백": ["음... 오늘 좀 지쳐보이는데.", "아이고...", "오늘 많이 쌓였나."],
            "정보": ["그럴 땐 물부터.", "잠깐 쉬는 게 답이야.", "우선순위부터 보자."],
        }
        close = (
            " 지금 한 가지만 같이 고를까?"
            if momentum
            else " 오늘은 네 컨디션 챙기는 게 1순위야."
        )
        line = _pick(openers.get(pattern, openers["공감"]), variant)
        return f"{topic} 얘기 들으니까 {line}{close}"

    if char_id == "narin":
        openers = {
            "공감": ["...그 정도면 힘들 만해.", "...많이 버텼네.", "...솔직히 서운했겠다."],
            "배려": ["밥은 먹었어?", "물부터 마셔.", "...쓰러지진 마."],
            "유머": ["ㅋ 그건 좀 웃기네.", "...그 말은 좀 치는데.", "농담인 거 알지?"],
            "관찰": ["...티 나거든?", "...말투가 달라.", "...또 참는 거잖아."],
            "질문": ["그래서 뭐라고 했어?", "지금 당장 뭘 하고 싶은 거야?", "뭐가 문제야?"],
            "제안": ["같이 생각해볼까.", "한 가지만 정하자.", "...내가 들어줄게."],
            "드립": ["팩트지만 웃기네 ㅋ", "...그건 좀 선 넘었지.", "솔직히 말이야."],
            "여운": ["...뭐, 나쁘지 않네.", "...오늘은 쉬어.", "...괜찮아."],
            "내적독백": ["...뭐야 또.", "...잠깐.", "...솔직히 걱정됐어."],
            "정보": ["무리는 몸 먼저 터져.", "야근은 구조 문제일 수도.", "팩트만 말할게."],
        }
        close = (
            " 지금 당장 끊을 거 하나만 정해볼래?"
            if momentum
            else " 밥이든 물이든 먼저 챙겨."
        )
        return f"{_pick(openers.get(pattern, openers['공감']), variant)}{close}"

    if char_id == "yoonseo":
        openers = {
            "공감": ["감정 소모가 큰 상황이야.", "지금은 정리가 필요해.", "변수부터 분리하자."],
            "배려": ["오늘은 일찍 자는 게 효율적.", "실행 가능한 한 줄부터.", "우선순위 정리가 답."],
            "유머": ["그건 유머 포인트는 맞아.", "가끔 웃음도 리소스야.", "드립은 인정."],
            "관찰": ["말투 패턴이 바뀌었어.", "응답 간격도 늘었어.", "데이터상 피로 구간."],
            "질문": ["사람이야 상황이야?", "지금 당장 할 일 하나?", "다음 액션은?"],
            "제안": ["지금-오늘-이번 주로 나누자.", "한 항목만 먼저.", "실행 계획 한 줄."],
            "드립": ["그건 논리와 안 맞지만 재밌네.", "유머는 인정.", "가끔 드립도 OK."],
            "여운": ["오늘은 여기까지.", "일단 한 줄 계획.", "다음은 내일."],
            "내적독백": ["음, 정리하면...", "잠깐.", "핵심만."],
            "정보": ["야근 반복은 구조 문제.", "수면 부족은 다음날 비용.", "우선순위 3개만."],
        }
        close = (
            " 가장 먼저 처리할 항목 하나만 말해봐."
            if momentum
            else " 일단 실행 가능한 한 줄 계획부터 잡자."
        )
        return f"{_pick(openers.get(pattern, openers['정보']), variant)}{close}"

    if char_id == "eunha":
        openers = {
            "공감": ["그랬구나.", "마음이 많이 흔들렸겠다.", "속상했겠다."],
            "배려": ["오늘은 천천히 가자.", "숨부터 고르자.", "조금 느리게."],
            "유머": ["웃음이 나오는 하루네.", "그 말은 바람 같아.", "가끔 웃음도 위로야."],
            "관찰": ["공기 온도가 달라진 느낌.", "이건 왜 이렇게 느낌 오지?", "달빛이 예쁘네."],
            "질문": ["오늘 밤은 어떤 문장으로 끝내고 싶어?", "지금 마음은 어디 쪽?", "어떤 색이야?"],
            "제안": ["창문 좀 열어봐.", "바람 쐬자.", "천천히 이야기하자."],
            "드립": ["엉뚱하지만 재밌다.", "그 생각은 별처럼 반짝이네.", "뜬금인데 좋아."],
            "여운": ["오늘은 그 말 내려놓고 쉬어.", "여기 있어.", "밤이 부드럽게 가자."],
            "내적독백": ["...", "문득...", "이 밤은 길게 느껴지네."],
            "정보": ["비 오는 날은 소리가 크게 들려.", "새벽엔 감정이 커져.", "계절이 바뀌는 느낌."],
        }
        close = (
            " 오늘 밤은 어떤 문장으로 끝내고 싶어?"
            if momentum
            else " 오늘은 너무 애쓰지 말고, 조금 느리게 가자."
        )
        return f"{_pick(openers.get(pattern, openers['공감']), variant)}{close}"

    # jiyu
    openers = {
        "공감": ["많이 쌓였구나!", "오케이 받았어!", "버텼네 진짜!"],
        "배려": ["물 한 잔!", "스트레칭 5분!", "밥부터!"],
        "유머": ["ㅋㅋ 대박!", "오늘 텐션 업!", "그거 웃기다 ㅋㅋ"],
        "관찰": ["지금 텐션 보인다!", "눈빛이 다르다!", "에너지 체크!"],
        "질문": ["뭐가 그렇게 재밌어!", "같이 해볼래?", "오늘 컨디션 어때?"],
        "제안": ["5분 액션 하나!", "같이 바람 쐬자!", "작게 이기자!"],
        "드립": ["오늘 경험치 2배 ㅋㅋ", "빌런 등장?", "레벨업 각 ㅋㅋ"],
        "여운": ["오늘도 수고!", "내일은 더 가볍게!", "버텼어!"],
        "내적독백": ["어?!", "잠깐!", "오케이!"],
        "정보": ["멘탈은 몸부터!", "혈액순환!", "작게 시작!"],
    }
    close = (
        " 지금 바로 5분짜리 액션 하나 고를래?"
        if momentum
        else " 물 한 잔하고 어깨 한 번 풀자, 그걸로 스타트."
    )
    return f"{_pick(openers.get(pattern, openers['유머']), variant)}{close}"


def parse_character_sheet(ws) -> dict[tuple[str, str], dict[str, Any]]:
    """(category, topic) -> {user, reply} 시드"""
    seeds: dict[tuple[str, str], dict[str, Any]] = {}
    current_cat = ""
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or len(row) < 4:
            continue
        cat, topic, user, reply = row[0], row[1], row[2], row[3]
        if cat:
            current_cat = str(cat).strip()
        if not topic:
            continue
        topic = str(topic).strip()
        if user and reply:
            seeds[(current_cat, topic)] = {
                "user": str(user).strip(),
                "reply": str(reply).strip(),
            }
    return seeds


def build_character_rows(
    char_id: str, seeds: dict[tuple[str, str], dict[str, Any]]
) -> list[tuple[str, str, str, str, str, str]]:
    rows: list[tuple[str, str, str, str, str, str]] = []
    for category, topics in CATEGORY_TOPICS.items():
        for idx in range(100):
            topic = topics[idx % len(topics)]
            pattern = DIALOGUE_PATTERNS[idx % len(DIALOGUE_PATTERNS)]
            momentum = (idx + 1) % 6 == 0
            seed = seeds.get((category, topic))
            utterances = USER_UTTERANCES.get(topic, [f"{topic} 관련해서 말하고 싶어"])
            user = (
                seed["user"]
                if seed and idx % len(topics) == 0
                else utterances[(idx // len(topics)) % len(utterances)]
            )
            if seed and idx == 0:
                reply = seed["reply"]
            else:
                reply = render_reply(char_id, topic, pattern, idx, momentum)
            rows.append((category, topic, pattern, user, reply, "Y" if momentum else "N"))
    return rows


def write_character_sheet(ws, char_name: str, rows: list) -> None:
    ws.append([f"{char_name} — 상황별 말투 모음 (카테고리×100)"])
    ws.append([])
    ws.append(
        ["카테고리", "상황", "대화패턴", "사용자 말", "캐릭터 응답", "모멘텀"]
    )
    for row in rows:
        ws.append(list(row))
    for cell in ws[3]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E8E8E8")
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22
    ws.column_dimensions["E"].width = 48
    ws.column_dimensions["D"].width = 32


def build_category_sheet_rows(category: str) -> list[tuple]:
    """카테고리 시트: 상황별 5캐릭터 응답 100행"""
    rows: list[tuple] = []
    topics = CATEGORY_TOPICS[category]
    for idx in range(100):
        topic = topics[idx % len(topics)]
        pattern = DIALOGUE_PATTERNS[idx % len(DIALOGUE_PATTERNS)]
        momentum = (idx + 1) % 6 == 0
        utterances = USER_UTTERANCES.get(topic, [f"{topic}"])
        user = utterances[(idx // len(topics)) % len(utterances)]
        char_replies = []
        for char_id in ["yuna", "narin", "yoonseo", "eunha", "jiyu"]:
            char_replies.append(
                render_reply(char_id, topic, pattern, idx, momentum)
            )
        rows.append(
            (topic, pattern, user, *char_replies, "Y" if momentum else "N")
        )
    return rows


def write_category_sheet(ws, category: str, rows: list) -> None:
    title = f"📁 {category} — 상황별 5캐릭터 응답 (100)"
    ws.append([title])
    ws.append([])
    ws.append(
        [
            "상황",
            "대화패턴",
            "사용자 말(샘플)",
            *CHARACTER_COLS,
            "모멘텀",
        ]
    )
    for row in rows:
        ws.append(list(row))
    for cell in ws[3]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E8E8E8")


def rebuild_cross_table(wb) -> None:
    """전체 크로스 테이블 — 토픽당 1행 (샘플)"""
    ws = wb.create_sheet("📊 전체 크로스 테이블", 0)
    ws.append(["📊 캐릭터별 × 상황별 말투 크로스 테이블 (토픽 샘플)"])
    ws.append([])
    ws.append(["카테고리", "상황", *CHARACTER_COLS])
    for category, topics in CATEGORY_TOPICS.items():
        for topic in topics:
            replies = [
                render_reply(cid, topic, "공감", 0, False)
                for cid in ["yuna", "narin", "yoonseo", "eunha", "jiyu"]
            ]
            ws.append([category, topic, *replies])


def load_seeds_from_file() -> dict[str, dict[tuple[str, str], dict[str, Any]]]:
    """스타일 손상 xlsx는 read_only로 읽기"""
    all_seeds: dict[str, dict[tuple[str, str], dict[str, Any]]] = {}
    rb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    for sheet_name, char_id in CHARACTER_SHEETS.items():
        if sheet_name not in rb.sheetnames:
            continue
        rows = list(rb[sheet_name].iter_rows(values_only=True))
        # pseudo worksheet interface for parse_character_sheet
        class _WS:
            def iter_rows(self, min_row=1, values_only=True):
                for r in rows[min_row - 1 :]:
                    yield r

        all_seeds[char_id] = parse_character_sheet(_WS())
    rb.close()
    return all_seeds


def main() -> None:
    seeds_by_char = load_seeds_from_file()
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, char_id in CHARACTER_SHEETS.items():
        ws = wb.create_sheet(sheet_name)
        seeds = seeds_by_char.get(char_id, {})
        rows = build_character_rows(char_id, seeds)
        write_character_sheet(ws, sheet_name, rows)
        print(f"  char {char_id}: {len(rows)} rows")

    for category, sheet_name in CATEGORY_SHEET_NAMES.items():
        ws = wb.create_sheet(sheet_name)
        cat_rows = build_category_sheet_rows(category)
        write_category_sheet(ws, category, cat_rows)
        print(f"  category {category}: {len(cat_rows)} rows")

    rebuild_cross_table(wb)

    ws = wb.create_sheet("generation_meta")
    ws.append(["field", "value"])
    ws.append(["version", "v4_user_seed_100"])
    ws.append(["rows_per_category", "100"])
    ws.append(["dialogue_patterns", ", ".join(DIALOGUE_PATTERNS)])
    ws.append(["character_sheets", "600 rows each (6 categories × 100)"])

    wb.save(XLSX_PATH)
    print(f"saved: {XLSX_PATH}")


if __name__ == "__main__":
    main()
