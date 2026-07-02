from __future__ import annotations

import re
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


CHARACTERS = ["유나", "나린", "윤서", "은하", "지유"]

CATEGORY_TOPICS: dict[str, list[str]] = {
    "감정 상태": ["짜증남", "무기력함", "외로움", "설렘", "뿌듯함", "불안함", "슬픔", "멍함"],
    "일상 이벤트": ["야근", "퇴근", "점심", "주말", "월요일 아침", "잠 못 잠", "운동함", "쉬는 날", "불금"],
    "관계/사람": ["친구에게 상처받음", "누군가 보고 싶음", "싸움", "칭찬받음", "무시당한 느낌", "반려동물의 죽음"],
    "시간대/날씨": ["새벽 감성", "비 오는 날", "첫눈", "더운 여름 낮", "명절", "대체공휴일"],
    "자기 자신": ["자신감 없음", "뭔가 이루고 싶음", "나태해진 느낌", "잘 해냈을 때", "머리가 복잡함"],
    "대화 행동(추가)": ["사과", "달래기", "축하", "칭찬", "주제 전환", "질문형 마무리", "여운형 마무리", "제안형 마무리"],
}

# 금지 표현 — 부분 일치 검사용 (정규식)
GLOBAL_FORBIDDEN_PATTERNS: list[tuple[str, str]] = [
    ("무슨 일 있어", r"무슨\s*일\s*있어"),
    ("괜찮아?", r"괜찮아\??"),
    ("~거 아니야", r"거\s*아니야\??"),
    ("도와드릴까요", r"도와\s*드릴까요"),
    ("(웃으며)", r"\(웃으며\)"),
    ("AI 언급", r"\bAI\b|인공지능|프로그램"),
]

CHARACTER_FORBIDDEN: dict[str, list[tuple[str, str]]] = {
    "유나": [("사랑해", r"사랑해")],
    "나린": [
        ("참나", r"참나"),
        ("네 탓", r"네\s*탓"),
        ("❤️", r"❤"),
    ],
    "윤서": [
        ("느낌표 과다", r"!{2,}"),
        ("보고 싶어", r"보고\s*싶어"),
    ],
    "은하": [("느낌표 과다", r"!{2,}")],
    "지유": [],
}

# big-kick 후보 토픽 (감정 피크·축하·새벽)
BIG_KICK_TOPICS = {
    "설렘",
    "뿌듯함",
    "축하",
    "잘 해냈을 때",
    "새벽 감성",
    "반려동물의 죽음",
    "슬픔",
}

USER_UTTERANCE_MAP: dict[str, list[str]] = {
    "짜증남": ["아 진짜 오늘 너무 짜증나", "하... 열받아서 말이 안 나와", "오늘 빌런 제대로 만났어"],
    "무기력함": ["요즘 아무것도 하기 싫어", "에너지가 바닥이야", "몸도 마음도 축 처진다"],
    "외로움": ["오늘따라 너무 외롭다", "괜히 마음이 비어있는 느낌이야", "사람 많은데도 혼자인 기분이야"],
    "설렘": ["오늘 좀 설레는 일이 생겼어", "괜히 기분이 들뜨네", "심장이 빨리 뛰는 느낌이야"],
    "뿌듯함": ["오늘 진짜 잘한 것 같아", "드디어 해냈다", "내가 좀 대견하다"],
    "불안함": ["괜히 불안해서 집중이 안 돼", "머릿속이 복잡해", "아직 안 일어난 일인데 걱정돼"],
    "슬픔": ["오늘 좀 많이 슬프네", "눈물 날 것 같다", "마음이 너무 가라앉아"],
    "멍함": ["멍하게 시간만 흘렀어", "오늘 정신이 하나도 없어", "뭘 했는지 기억이 안 나"],
    "야근": ["오늘 또 야근 확정이야", "퇴근이 안 보인다", "지금도 회사야"],
    "퇴근": ["이제 퇴근한다", "오늘도 겨우 버텼네", "집 가는 길이야"],
    "점심": ["점심 뭐 먹지", "배는 고픈데 귀찮다", "오늘 점심 추천해줘"],
    "주말": ["주말인데 뭘 해야 할지 모르겠어", "이번 주말은 쉬고 싶다", "주말 계획 아직 없음"],
    "월요일 아침": ["월요일 아침부터 힘들다", "출근길부터 텐션 바닥", "월요병 제대로 왔어"],
    "잠 못 잠": ["어제 잠을 거의 못 잤어", "새벽까지 뒤척였다", "눈은 감았는데 머리가 안 쉬더라"],
    "운동함": ["오늘 운동하고 왔어", "땀 좀 빼니까 살 것 같다", "오랜만에 몸 풀었다"],
    "쉬는 날": ["오늘은 쉬는 날이야", "아무 일정 없이 쉬고 있다", "오늘은 집콕 모드"],
    "불금": ["불금인데 뭐하지", "오늘은 좀 놀고 싶다", "불금 텐션 올려줘"],
    "친구에게 상처받음": ["친구한테 말로 상처받았어", "가까운 사람한테 실망했다", "친구 때문에 마음이 아프다"],
    "누군가 보고 싶음": ["갑자기 누가 보고 싶다", "문득 생각나는 사람이 있어", "그 사람이 자꾸 떠오른다"],
    "싸움": ["아까 크게 다퉜어", "말하다가 싸워버렸어", "감정이 너무 올라왔어"],
    "칭찬받음": ["오늘 칭찬받았어", "생각보다 좋은 평가 받았다", "인정받으니까 기분 좋다"],
    "무시당한 느낌": ["오늘 무시당한 기분 들었어", "내 말이 투명해진 느낌이야", "존중 못 받은 느낌이라 별로다"],
    "반려동물의 죽음": ["반려동물이 떠났어", "마음이 텅 빈 것 같다", "아직 실감이 안 난다"],
    "새벽 감성": ["새벽이라 그런지 생각이 많아", "이 시간엔 감정이 커진다", "잠도 안 오고 마음이 복잡해"],
    "비 오는 날": ["비 오는 소리 들으니까 기분이 묘해", "오늘 하루 종일 비네", "비 오니까 괜히 울적해"],
    "첫눈": ["오늘 첫눈 봤어", "창밖에 눈 오는데 예쁘더라", "첫눈 보니까 마음이 몽글해"],
    "더운 여름 낮": ["너무 더워서 녹을 것 같다", "여름 낮 진짜 버티기 힘들다", "오늘 햇빛이 장난 아니야"],
    "명절": ["명절이라 정신없다", "가족 모임이 좀 버겁다", "명절 분위기가 묘하게 복잡해"],
    "대체공휴일": ["대체공휴일인데 쉬는 게 어색하다", "하루 번 느낌이라 좋다", "오늘은 느긋하게 보내고 싶어"],
    "자신감 없음": ["요즘 자신감이 너무 떨어진다", "내가 잘하고 있는지 모르겠어", "괜히 다 작아 보인다"],
    "뭔가 이루고 싶음": ["이번엔 진짜 뭐 하나 이루고 싶어", "변화가 필요해", "목표를 제대로 잡고 싶다"],
    "나태해진 느낌": ["요즘 너무 늘어진 느낌이야", "해야 할 걸 자꾸 미룬다", "리듬이 완전 깨졌다"],
    "잘 해냈을 때": ["이번 건 진짜 잘 해낸 것 같아", "끝내고 나니까 뿌듯하다", "고생한 보람이 있네"],
    "머리가 복잡함": ["머릿속이 너무 시끄러워", "정리해도 정리가 안 된다", "해야 할 게 많아서 멘붕이야"],
    "사과": ["아까 말 심했지 미안해", "내가 먼저 사과할게", "그건 내 실수였다"],
    "달래기": ["오늘 위로 좀 해줘", "따뜻한 말이 필요해", "마음이 너무 예민하다"],
    "축하": ["좋은 소식 있어", "이번에 진짜 잘됐다", "드디어 목표 달성했어"],
    "칭찬": ["나 칭찬 좀 해줘", "오늘 나 괜찮지?", "잘한 거 맞지?"],
    "주제 전환": ["분위기 바꿔볼까", "이 얘기는 여기까지 할래", "다른 얘기 하고 싶어"],
    "질문형 마무리": ["짧게 마무리하고 싶어", "한마디만 듣고 싶다", "오늘은 짧게 얘기하자"],
    "여운형 마무리": ["조용하게 끝내고 싶어", "오늘은 잔잔하게 마무리하고 싶다", "여운 남는 말 한 줄 해줘"],
    "제안형 마무리": ["다음 액션 하나만 정해줘", "내일 할 거 한 개만 정하자", "작게라도 같이 해보자"],
}


def render_reply(character: str, topic: str, momentum: bool, variant: int) -> str:
    if character == "유나":
        base = [
            f"오늘 {topic} 때문에 마음이 많이 흔들렸겠다. 네 편에서 같이 정리해볼게.",
            f"{topic} 얘기 듣는데 네가 얼마나 버텼는지 느껴져. 일단 숨 좀 고르자.",
        ][variant % 2]
        return base + (" 지금 제일 급한 한 가지만 같이 고를까?" if momentum else " 오늘은 네 컨디션 챙기는 게 1순위야.")

    if character == "나린":
        base = [
            f"...{topic}면 솔직히 빡셀 만하지. 근데 너 또 혼자 다 버틴 거잖아.",
            f"{topic} 상황에서 그 정도면 충분히 잘한 거야. 무리하는 건 멋있는 게 아니고.",
        ][variant % 2]
        return base + (" 그래서 지금 당장 끊어낼 거 하나만 정해볼래?" if momentum else " 밥이든 물이든 먼저 챙겨.")

    if character == "윤서":
        base = [
            f"{topic} 이슈는 감정 소모가 큰 편이야. 먼저 변수부터 분리하자.",
            f"{topic} 상황이면 우선순위 정리가 답이야. 지금-오늘-이번 주로 나누면 된다.",
        ][variant % 2]
        return base + (" 지금 기준으로 가장 먼저 처리할 항목 하나만 말해봐." if momentum else " 일단 실행 가능한 한 줄 계획부터 잡자.")

    if character == "은하":
        base = [
            f"{topic}의 결은 유독 마음을 오래 붙잡아 두지. 그래도 네 호흡은 아직 따뜻해.",
            f"오늘 {topic} 얘기를 들으니까 공기 온도가 달라진 느낌이야. 너의 마음도 그럴 거고.",
        ][variant % 2]
        return base + (" 오늘 밤은 어떤 문장으로 끝내고 싶어?" if momentum else " 오늘은 너무 애쓰지 말고, 조금 느리게 가자.")

    base = [
        f"{topic} 모드면 멘탈 체력부터 챙겨야 해 ㅋㅋ 너 생각보다 잘 버티는 타입이야.",
        f"오케이 {topic} 접수! 오늘은 작게 이기면 되는 날이다.",
    ][variant % 2]
    return base + (" 지금 바로 할 수 있는 5분짜리 액션 하나 고를래?" if momentum else " 물 한 잔하고 어깨 한 번 풀자, 그걸로 스타트.")


def count_ellipsis(text: str) -> int:
    return len(re.findall(r"\.{3,}|…+", text))


def detect_forbidden_hits(character: str, text: str) -> list[str]:
    hits: list[str] = []
    patterns = GLOBAL_FORBIDDEN_PATTERNS + CHARACTER_FORBIDDEN.get(character, [])
    for label, pattern in patterns:
        if re.search(pattern, text, re.IGNORECASE):
            hits.append(label)
    if count_ellipsis(text) > 1:
        hits.append("말줄임 2회 이상")
    if re.fullmatch(r"[ㅋㅎ]+", text.strip()):
        hits.append("ㅋㅋ 단독")
    return hits


def classify_kick(
    character: str, topic: str, momentum: bool, variant: int
) -> tuple[str, str]:
    """kick_candidate(Y/N), kick_type(none|micro-kick|big-kick)"""
    if not momentum:
        return "N", "none"

    # big-kick: 희소 — 감정 피크 토픽 + 30턴 주기
    if topic in BIG_KICK_TOPICS and (variant + 1) % 30 == 0:
        if character == "윤서" and topic == "새벽 감성":
            return "Y", "big-kick"
        if character != "윤서" and topic in BIG_KICK_TOPICS:
            return "Y", "big-kick"

    # micro-kick: 모멘텀 턴의 약 35%
    if (variant + 1) % 3 == 0:
        return "Y", "micro-kick"

    return "N", "none"


def build_pair_rows() -> Iterable[
    tuple[str, str, str, str, str, str, str, str, str, str]
]:
    for category, topics in CATEGORY_TOPICS.items():
        for idx in range(100):
            topic = topics[idx % len(topics)]
            character = CHARACTERS[idx % len(CHARACTERS)]
            utterances = USER_UTTERANCE_MAP[topic]
            user_utterance = utterances[idx % len(utterances)]
            momentum = (idx + 1) % 6 == 0
            assistant = render_reply(character, topic, momentum, idx)
            kick_candidate, kick_type = classify_kick(
                character, topic, momentum, idx
            )
            forbidden_hits = detect_forbidden_hits(character, assistant)
            forbidden_hit = "Y" if forbidden_hits else "N"
            qa_pass = "PASS" if forbidden_hit == "N" else "FAIL"
            yield (
                category,
                topic,
                character,
                user_utterance,
                assistant,
                "Y" if momentum else "N",
                kick_candidate,
                kick_type,
                forbidden_hit,
                "; ".join(forbidden_hits) if forbidden_hits else "",
                qa_pass,
            )


def autosize_columns(ws) -> None:
    for col_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(
            max(length + 2, 12), 78
        )


def highlight_qa_failures(ws) -> None:
    fail_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    header = [c.value for c in ws[1]]
    qa_col = header.index("qa_pass") + 1
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if row[qa_col - 1].value == "FAIL":
            for cell in row:
                cell.fill = fail_fill


def build_qa_summary(rows: list[tuple]) -> list[tuple[str, str | int]]:
    total = len(rows)
    fail = sum(1 for r in rows if r[-1] == "FAIL")
    kick_y = sum(1 for r in rows if r[6] == "Y")
    micro = sum(1 for r in rows if r[7] == "micro-kick")
    big = sum(1 for r in rows if r[7] == "big-kick")
    momentum = sum(1 for r in rows if r[5] == "Y")
    return [
        ("total_rows", total),
        ("qa_pass", total - fail),
        ("qa_fail", fail),
        ("momentum_turns", momentum),
        ("kick_candidate_y", kick_y),
        ("micro_kick", micro),
        ("big_kick", big),
        ("pass_rate_pct", round((total - fail) / total * 100, 1) if total else 0),
    ]


def main() -> None:
    out_dir = Path("data/dialogue_examples")
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "character_dialogue_examples.xlsx"

    rows = list(build_pair_rows())

    wb = Workbook()

    ws_pairs = wb.active
    ws_pairs.title = "dialogue_pairs_100"
    ws_pairs.append(
        [
            "category",
            "topic",
            "character",
            "user_utterance",
            "assistant_reply",
            "momentum_turn",
            "kick_candidate",
            "kick_type",
            "forbidden_hit",
            "forbidden_detail",
            "qa_pass",
        ]
    )
    for row in rows:
        ws_pairs.append(list(row))
    highlight_qa_failures(ws_pairs)
    autosize_columns(ws_pairs)

    ws_qa = wb.create_sheet("qa_summary")
    ws_qa.append(["metric", "value"])
    for metric, value in build_qa_summary(rows):
        ws_qa.append([metric, value])
    autosize_columns(ws_qa)

    ws_forbidden = wb.create_sheet("forbidden_rules")
    ws_forbidden.append(["scope", "label", "pattern"])
    for label, pattern in GLOBAL_FORBIDDEN_PATTERNS:
        ws_forbidden.append(["global", label, pattern])
    for char, patterns in CHARACTER_FORBIDDEN.items():
        for label, pattern in patterns:
            ws_forbidden.append([char, label, pattern])
    autosize_columns(ws_forbidden)

    ws_style = wb.create_sheet("style_ratio_note")
    ws_style.append(
        ["character", "base_ratio", "momentum_ratio", "internal_monologue_ratio", "note"]
    )
    ws_style.append(["유나", "80", "20", "10%", "편안함/안정감 중심"])
    ws_style.append(["나린", "80", "20", "15%", "다정함 우선 + 가벼운 팩트"])
    ws_style.append(["윤서", "80", "20", "8%", "정리/판단 중심"])
    ws_style.append(["은하", "80", "20", "20%", "감성/상상 확장"])
    ws_style.append(["지유", "80", "20", "5%", "텐션/드립"])
    autosize_columns(ws_style)

    ws_meta = wb.create_sheet("generation_meta")
    ws_meta.append(["field", "value"])
    ws_meta.append(["version", "v3_qa"])
    ws_meta.append(["rows_per_category", "100"])
    ws_meta.append(["momentum_rule", "every 6 turns"])
    ws_meta.append(["kick_micro", "momentum + variant%3==0"])
    ws_meta.append(["kick_big", "momentum + peak topic + variant%30==0"])
    ws_meta.append(["qa_columns", "kick_candidate,kick_type,forbidden_hit,qa_pass"])
    autosize_columns(ws_meta)

    wb.save(out_path)
    summary = build_qa_summary(rows)
    print(f"generated: {out_path}")
    for k, v in summary:
        print(f"  {k}: {v}")


if __name__ == "__main__":
    main()
