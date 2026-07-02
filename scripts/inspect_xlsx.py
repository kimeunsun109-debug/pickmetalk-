import sys
from pathlib import Path
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")
p = Path("character_dialogue_examples.xlsx")
wb = load_workbook(p, read_only=True)
for name in wb.sheetnames:
    ws = wb[name]
    print(f"=== {name!r} rows={ws.max_row} cols={ws.max_column}")
    rows = list(ws.iter_rows(max_row=8, values_only=True))
    for r in rows:
        print(r)
    print()
wb.close()
